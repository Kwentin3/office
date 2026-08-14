from __future__ import annotations

import copy
import json
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest import mock

from lxml import etree
from openpyxl import Workbook, load_workbook
from xlsx_artifact_tool import XlsxArtifactTool


class StrictTemplateTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.tool = XlsxArtifactTool(self.root / "work")
        self.source = self.root / "template.xlsx"
        workbook = Workbook()
        workbook.active.title = "Offer"
        workbook.active["A1"] = "Customer: {{customer}}"
        workbook.active["B2"] = "{{amount}} {{currency}}"
        workbook.create_sheet("Meta")["A1"] = "Reference {{reference}}"
        workbook.save(self.source)

    def tearDown(self) -> None:
        self.temp.cleanup()

    def test_strict_fill_resolves_string_cells_across_sheets(self) -> None:
        output = self.root / "filled.xlsx"
        result = self.tool.fill_template(
            self.source,
            {"customer": "Acme", "amount": "100", "currency": "USD", "reference": "R-42"},
            output,
        )
        self.assertEqual(result["status"], "ok")
        workbook = load_workbook(output, data_only=False)
        self.assertEqual(workbook["Offer"]["A1"].value, "Customer: Acme")
        self.assertEqual(workbook["Offer"]["B2"].value, "100 USD")
        self.assertEqual(workbook["Meta"]["A1"].value, "Reference R-42")

    def test_strict_fill_refuses_missing_or_unknown_values(self) -> None:
        for values in (
            {"customer": "Acme"},
            {"customer": "Acme", "amount": "100", "currency": "USD", "reference": "R-42", "extra": "x"},
        ):
            with self.subTest(values=values):
                output = self.root / "never.xlsx"
                result = self.tool.fill_template(self.source, values, output)
                self.assertEqual(result["status"], "refused")
                self.assertEqual(result["reason"], "validation_failure")
                self.assertFalse(output.exists())

    def test_strict_fill_rejects_package_wide_marker(self) -> None:
        rewritten = self.root / "package-marker.xlsx"
        with zipfile.ZipFile(self.source) as package, zipfile.ZipFile(rewritten, "w") as target:
            for info in package.infolist():
                payload = package.read(info.filename)
                if info.filename == "docProps/core.xml":
                    root = etree.fromstring(payload)
                    root.set("{urn:kwentin:test}marker", "{{residual}}")
                    payload = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)
                target.writestr(copy.copy(info), payload)
        output = self.root / "package-marker-output.xlsx"
        result = self.tool.fill_template(
            rewritten,
            {"customer": "Acme", "amount": "100", "currency": "USD", "reference": "R-42"},
            output,
        )
        self.assertEqual((result["status"], result["reason"]), ("refused", "unsupported_capability"))
        self.assertFalse(output.exists())

    def test_strict_fill_refuses_malformed_overflow_formula_and_header_scope(self) -> None:
        cases = []

        malformed = Workbook()
        malformed.active["A1"] = "{{{name}}}"
        cases.append((malformed, {"name": "Acme"}, "validation_failure"))

        overflow = Workbook()
        overflow.active["A1"] = "X{{name}}"
        cases.append((overflow, {"name": "x" * 32767}, "validation_failure"))

        formula = Workbook()
        formula.active["A1"] = "={{name}}"
        cases.append((formula, {"name": "A2"}, "unsupported_capability"))

        header = Workbook()
        header.active["A1"] = "{{name}}"
        header.active.oddHeader.center.text = "Hidden {{header}}"
        cases.append((header, {"name": "Acme", "header": "x"}, "unsupported_capability"))

        for index, (workbook, values, reason) in enumerate(cases):
            with self.subTest(index=index, reason=reason):
                source = self.root / f"unsafe-{index}.xlsx"
                output = self.root / f"unsafe-output-{index}.xlsx"
                workbook.save(source)
                result = self.tool.fill_template(source, values, output)
                self.assertEqual((result["status"], result["reason"]), ("refused", reason))
                self.assertFalse(output.exists())

    def test_strict_fill_rejects_dtd_entity_declarations_package_wide(self) -> None:
        declaration = '<!DOCTYPE cp:coreProperties [<!ENTITY hidden "{{hidden}}">]>'
        for encoding in ("UTF-8", "UTF-16"):
            with self.subTest(encoding=encoding):
                rewritten = self.root / f"dtd-marker-{encoding}.xlsx"
                with zipfile.ZipFile(self.source) as package, zipfile.ZipFile(rewritten, "w") as target:
                    for info in package.infolist():
                        payload = package.read(info.filename)
                        if info.filename == "docProps/core.xml":
                            root = etree.fromstring(payload)
                            payload = etree.tostring(
                                root,
                                xml_declaration=True,
                                encoding=encoding,
                                doctype=declaration,
                            )
                        target.writestr(copy.copy(info), payload)
                output = self.root / f"dtd-output-{encoding}.xlsx"
                result = self.tool.fill_template(
                    rewritten,
                    {"customer": "Acme", "amount": "100", "currency": "USD", "reference": "R-42"},
                    output,
                )
                self.assertEqual(result["status"], "refused")
                self.assertFalse(output.exists())

    def test_strict_fill_preserves_unknown_package_members(self) -> None:
        rewritten = self.root / "custom-member.xlsx"
        payload = b"<business>KEEP</business>"
        with zipfile.ZipFile(self.source) as package, zipfile.ZipFile(rewritten, "w") as target:
            for info in package.infolist():
                target.writestr(copy.copy(info), package.read(info.filename))
            target.writestr("customXml/business.xml", payload)
        output = self.root / "custom-member-output.xlsx"
        result = self.tool.fill_template(
            rewritten,
            {"customer": "Acme", "amount": "100", "currency": "USD", "reference": "R-42"},
            output,
        )
        self.assertEqual(result["status"], "ok")
        with zipfile.ZipFile(output) as package:
            self.assertEqual(package.read("customXml/business.xml"), payload)

    def test_cli_fill_template_matches_api_contract(self) -> None:
        output = self.root / "cli.xlsx"
        payload = {
            "action": "fill_template",
            "workdir": str(self.root / "cli-work"),
            "source": str(self.source),
            "values": {"customer": "Acme", "amount": "100", "currency": "USD", "reference": "R-42"},
            "output": str(output),
        }
        completed = subprocess.run(
            [sys.executable, "-m", "xlsx_artifact_tool"],
            input=json.dumps(payload),
            text=True,
            capture_output=True,
            check=False,
        )
        self.assertEqual(completed.returncode, 0, completed.stderr)
        self.assertEqual(json.loads(completed.stdout)["status"], "ok")
        self.assertTrue(output.is_file())

    def test_strict_fill_rejects_residual_marker_in_unsupported_candidate_scope(self) -> None:
        source = self.root / "residual-source.xlsx"
        workbook = Workbook()
        workbook.active["A1"] = "{{name}}"
        workbook.save(source)
        output = self.root / "residual-output.xlsx"
        original_apply = self.tool.apply

        def inject_residual(source_path, plan, candidate):
            result = original_apply(source_path, plan, candidate)
            if result.get("status") == "ok":
                rewritten = self.root / "residual-candidate.xlsx"
                with zipfile.ZipFile(candidate) as package, zipfile.ZipFile(rewritten, "w") as target:
                    for info in package.infolist():
                        payload = package.read(info.filename)
                        if info.filename == "docProps/core.xml":
                            root = etree.fromstring(payload)
                            root.set("{urn:kwentin:test}residual", "{{residual}}")
                            payload = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)
                        target.writestr(copy.copy(info), payload)
                rewritten.replace(candidate)
            return result

        with mock.patch.object(self.tool, "apply", side_effect=inject_residual):
            result = self.tool.fill_template(source, {"name": "Acme"}, output)
        self.assertEqual((result["status"], result["reason"]), ("refused", "validation_failure"))
        self.assertFalse(output.exists())


if __name__ == "__main__":
    unittest.main()
