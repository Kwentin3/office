from __future__ import annotations
import copy,tempfile,unittest,zipfile
from pathlib import Path
from unittest.mock import patch
from openpyxl import load_workbook
from openpyxl.comments import Comment
from xlsx_artifact_tool import XlsxArtifactTool
from xlsx_artifact_tool.api import _object_sha

class SafetyTests(unittest.TestCase):
 def setUp(self):
  self.temp=tempfile.TemporaryDirectory();self.root=Path(self.temp.name);self.tool=XlsxArtifactTool(self.root/'work');self.source=self.root/'source.xlsx';self.tool.create({'sheets':[{'name':'Data','cells':{'A1':{'value':'Item','style':'header'},'A2':{'value':'A'},'B2':{'value':2}}}]},self.source)
 def tearDown(self):self.temp.cleanup()
 def test_stale_forged_malformed_and_duplicate_writes_are_refused_without_output(self):
  snap=self.tool.inspect(self.source,view='region',sheet='Data',range_ref='A1:B2');target=next(x for x in snap['cells'] if x['coordinate']=='B2');planned=self.tool.plan(snap,{'operations':[{'type':'set_cell_value','target_id':target['id'],'value':3,'expected_kind':'value'}]})['plan']
  forged=copy.deepcopy(planned);forged['operations']=forged['operations']*1001;raw=dict(forged);raw.pop('plan_sha256');forged['plan_sha256']=_object_sha(raw);out=self.root/'forged.xlsx';result=self.tool.apply(self.source,forged,out);self.assertEqual(result['status'],'refused');self.assertEqual(result['reason'],'unsafe_plan');self.assertFalse(out.exists())
  duplicate=copy.deepcopy(planned);duplicate['operations']=duplicate['operations']*2;raw=dict(duplicate);raw.pop('plan_sha256');duplicate['plan_sha256']=_object_sha(raw);result=self.tool.apply(self.source,duplicate,out);self.assertEqual(result['status'],'refused');self.assertEqual(result['reason'],'conflict');self.assertFalse(out.exists())
  changed=self.root/'changed.xlsx';self.tool.create({'sheets':[{'name':'Data','cells':{'A1':{'value':'Other'}}}]},changed);result=self.tool.apply(changed,planned,out);self.assertEqual(result['reason'],'stale_snapshot');self.assertFalse(out.exists())
  malformed=self.tool.plan(snap,{'operations':[{'type':'set_cell_value','target_id':target['id'],'value':object(),'expected_kind':'value'}]});self.assertEqual(malformed['status'],'refused')
 def test_value_edit_preserves_comment_and_unrelated_package_members(self):
  decorated=self.root/'decorated-real.xlsx';wb=load_workbook(self.source);wb['Data']['A1'].comment=Comment('keep me','tester');wb.save(decorated)
  with zipfile.ZipFile(decorated) as archive:before={info.filename:archive.read(info.filename) for info in archive.infolist()}
  snap=self.tool.inspect(decorated,view='region',sheet='Data',range_ref='A1:B2');target=next(x for x in snap['cells'] if x['coordinate']=='B2');plan=self.tool.plan(snap,{'operations':[{'type':'set_cell_value','target_id':target['id'],'value':3,'expected_kind':'value'}]})['plan'];out=self.root/'preserved.xlsx'
  result=self.tool.apply(decorated,plan,out);self.assertEqual(result['status'],'ok')
  after=load_workbook(out);self.assertEqual(after['Data']['B2'].value,3);self.assertEqual(after['Data']['A1'].comment.text,'keep me')
  with zipfile.ZipFile(out) as archive:after_members={info.filename:archive.read(info.filename) for info in archive.infolist()}
  changed=[name for name in before.keys()|after_members.keys() if before.get(name)!=after_members.get(name)]
  self.assertEqual(changed,['xl/worksheets/sheet1.xml'])

 def test_formula_edit_preserves_comment_and_unrelated_package_members(self):
  decorated=self.root/'formula-decorated.xlsx';wb=load_workbook(self.source);wb['Data']['A1'].comment=Comment('keep formula book','tester');wb.save(decorated)
  with zipfile.ZipFile(decorated) as archive:before={info.filename:archive.read(info.filename) for info in archive.infolist()}
  snap=self.tool.inspect(decorated,view='region',sheet='Data',range_ref='A1:B2');target=next(x for x in snap['cells'] if x['coordinate']=='B2');plan=self.tool.plan(snap,{'operations':[{'type':'set_cell_formula','target_id':target['id'],'formula':'=1+2','expected_kind':'value'}]})['plan'];out=self.root/'formula-preserved.xlsx'
  result=self.tool.apply(decorated,plan,out);self.assertEqual(result['status'],'ok');self.assertEqual(result['formula_recalculation'],'required')
  after=load_workbook(out,data_only=False);self.assertEqual(after['Data']['B2'].value,'=1+2');self.assertEqual(after['Data']['A1'].comment.text,'keep formula book')
  with zipfile.ZipFile(out) as archive:after_members={info.filename:archive.read(info.filename) for info in archive.infolist()}
  changed=sorted(name for name in before.keys()|after_members.keys() if before.get(name)!=after_members.get(name))
  self.assertEqual(changed,['xl/worksheets/sheet1.xml'])

 def test_digital_signature_graph_is_refused_before_mutation(self):
  signed=self.root/'signed.xlsx'
  with zipfile.ZipFile(self.source) as src,zipfile.ZipFile(signed,'w') as dst:
   for info in src.infolist():dst.writestr(info,src.read(info.filename))
   dst.writestr('_xmlsignatures/origin.sigs',b'<origin/>');dst.writestr('_xmlsignatures/_rels/origin.sigs.rels',b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>');dst.writestr('_xmlsignatures/sig1.xml',b'<Signature/>')
  snap=self.tool.inspect(signed,view='region',sheet='Data',range_ref='A1:B2');target=next(x for x in snap['cells'] if x['coordinate']=='B2');plan=self.tool.plan(snap,{'operations':[{'type':'set_cell_value','target_id':target['id'],'value':3,'expected_kind':'value'}]})['plan'];out=self.root/'must-not-invalidate-signature.xlsx';result=self.tool.apply(signed,plan,out);self.assertEqual((result['status'],result['reason']),('refused','unsupported_capability'));self.assertFalse(out.exists())
 def test_apply_refuses_unsupported_package_features_before_candidate(self):
  decorated=self.root/'decorated.xlsx';decorated.write_bytes(self.source.read_bytes())
  rewritten=self.root/'rewritten.xlsx'
  with zipfile.ZipFile(decorated) as src,zipfile.ZipFile(rewritten,'w') as dst:
   for info in src.infolist():dst.writestr(info,src.read(info.filename))
   dst.writestr('xl/comments1.xml','<comments/>')
  rewritten.replace(decorated)
  snap=self.tool.inspect(decorated,view='region',sheet='Data',range_ref='A1:B2');self.assertEqual(snap['status'],'ok')
  plan=self.tool.plan(snap,{'operations':[{'type':'append_rows','region_id':snap['region_id'],'copy_from_row_id':snap['rows'][1]['id'],'rows':[['B',3]]}]})['plan'];out=self.root/'must-not-publish.xlsx'
  result=self.tool.apply(decorated,plan,out);self.assertEqual(result['status'],'refused');self.assertEqual(result['reason'],'unsupported_capability');self.assertFalse(out.exists())

 def test_formula_mode_refuses_dde_command_formulas(self):
  snap=self.tool.inspect(self.source,view='region',sheet='Data',range_ref='A1:B2');target=next(x for x in snap['cells'] if x['coordinate']=='B2')
  for formula in ("=cmd|' /C calc'!A0","=powershell|'x'!A0","=DDEAUTO('cmd','/c calc','A0')",r"=MSEXCEL|'C:\evil.xlsx'!R1C1"):
   result=self.tool.plan(snap,{'operations':[{'type':'set_cell_formula','target_id':target['id'],'formula':formula,'expected_kind':'value'}]});self.assertEqual(result['status'],'refused')
 def test_external_formula_and_unsafe_packages_are_refused(self):
  snap=self.tool.inspect(self.source,view='region',sheet='Data',range_ref='A1:B2');target=next(x for x in snap['cells'] if x['coordinate']=='B2');result=self.tool.plan(snap,{'operations':[{'type':'set_cell_formula','target_id':target['id'],'formula':"='[other.xlsx]Sheet1'!A1",'expected_kind':'value'}]});self.assertEqual(result['status'],'refused')
  unsafe=self.root/'unsafe.xlsx'
  with zipfile.ZipFile(unsafe,'w') as z:z.writestr('../escape.xml','x')
  self.assertEqual(self.tool.inspect(unsafe,view='summary')['status'],'refused');self.assertEqual(self.tool.validate(unsafe)['status'],'invalid')
 def test_missing_required_opc_member_is_refused(self):
  for index,missing in enumerate(('[Content_Types].xml','_rels/.rels','xl/workbook.xml','xl/_rels/workbook.xml.rels')):
   broken=self.root/f'missing-{index}.xlsx'
   with zipfile.ZipFile(self.source) as src,zipfile.ZipFile(broken,'w') as dst:
    for info in src.infolist():
     if info.filename!=missing:dst.writestr(info,src.read(info.filename))
   self.assertEqual(self.tool.inspect(broken,view='summary')['status'],'refused')
   self.assertEqual(self.tool.validate(broken)['status'],'invalid')
 def test_apply_is_bound_to_private_source_snapshot(self):
  snap=self.tool.inspect(self.source,view='region',sheet='Data',range_ref='A1:B2');target=next(x for x in snap['cells'] if x['coordinate']=='B2');plan=self.tool.plan(snap,{'operations':[{'type':'set_cell_value','target_id':target['id'],'value':3,'expected_kind':'value'}]})['plan'];out=self.root/'race-safe.xlsx'
  from xlsx_artifact_tool import api as api_module
  real_sha=api_module._sha;raced={'done':False}
  def race_after_hash(path):
   digest=real_sha(path)
   if Path(path)==self.source and not raced['done']:
    raced['done']=True;wb=load_workbook(self.source);wb['Data']['A1']='RACED';wb.save(self.source)
   return digest
  with patch('xlsx_artifact_tool.api._sha',side_effect=race_after_hash):result=self.tool.apply(self.source,plan,out)
  self.assertEqual(result['status'],'ok');self.assertTrue(out.exists());self.assertEqual(load_workbook(out)['Data']['A1'].value,'Item')
 def test_plan_and_apply_malformed_xml_are_typed_refusals(self):
  malformed_snapshot={'status':'ok','rows':None};result=self.tool.plan(malformed_snapshot,{'operations':[{}]});self.assertEqual((result['status'],result['reason']),('refused','validation_failure'))
  broken=self.root/'broken-workbook.xlsx'
  with zipfile.ZipFile(self.source) as src,zipfile.ZipFile(broken,'w') as dst:
   for info in src.infolist():dst.writestr(info,b'<broken' if info.filename=='xl/workbook.xml' else src.read(info.filename))
  plan={'schema':1,'source_sha256':__import__('hashlib').sha256(broken.read_bytes()).hexdigest(),'snapshot_sha256':'0'*64,'operations':[{'type':'set_cell_value'}]};plan['plan_sha256']=_object_sha(plan);out=self.root/'must-not-publish-broken.xlsx';result=self.tool.apply(broken,plan,out);self.assertEqual((result['status'],result['reason']),('refused','validation_failure'));self.assertFalse(out.exists())
 def test_public_xlsx_boundaries_never_leak_library_or_os_errors(self):
  malformed=self.root/'malformed.xlsx'
  with zipfile.ZipFile(self.source) as src,zipfile.ZipFile(malformed,'w') as dst:
   for info in src.infolist():dst.writestr(info,b'<broken' if info.filename=='xl/workbook.xml' else src.read(info.filename))
  result=self.tool.inspect(malformed,view='summary');self.assertEqual(result['status'],'refused')
  output_directory=self.root/'directory.xlsx';output_directory.mkdir()
  result=self.tool.create({'sheets':[{'name':'Data','cells':{}}]},output_directory);self.assertEqual(result['status'],'refused');self.assertTrue(output_directory.is_dir())
 def test_source_equals_output_and_postcheck_failure_never_publish(self):
  snap=self.tool.inspect(self.source,view='region',sheet='Data',range_ref='A1:B2');target=next(x for x in snap['cells'] if x['coordinate']=='B2');plan=self.tool.plan(snap,{'operations':[{'type':'set_cell_value','target_id':target['id'],'value':3,'expected_kind':'value'}]})['plan'];before=self.source.read_bytes();result=self.tool.apply(self.source,plan,self.source);self.assertEqual(result['reason'],'unsafe_plan');self.assertEqual(self.source.read_bytes(),before)

 def test_public_create_is_closed_and_returns_typed_refusals(self):
  bad_models=[{}, {'sheets':[{'name':'Data','cells':{},'unknown':1}]}, {'sheets':[{'name':'Data','cells':{}},{'name':'Data','cells':{}}]}, {'sheets':[{'name':'Bad/Name','cells':{}}]}, {'sheets':[{'name':'Data','cells':{'A1':{'value':1,'extra':2}}}]}, {'sheets':[{'name':'Data','cells':{'A1':{'value':'=1+2'}}}]}]
  for index,model in enumerate(bad_models):
   out=self.root/f'bad-{index}.xlsx';result=self.tool.create(model,out);self.assertEqual(result['status'],'refused');self.assertFalse(out.exists())
 def test_explicit_scalar_mode_refuses_formula_like_strings(self):
  snap=self.tool.inspect(self.source,view='region',sheet='Data',range_ref='A1:B2');target=next(x for x in snap['cells'] if x['coordinate']=='B2')
  direct=self.tool.plan(snap,{'operations':[{'type':'set_cell_value','target_id':target['id'],'value':'=1+2','expected_kind':'value'}]});self.assertEqual((direct['status'],direct['reason']),('refused','validation_failure'))
  appended=self.tool.plan(snap,{'operations':[{'type':'append_rows','region_id':snap['region_id'],'copy_from_row_id':snap['rows'][1]['id'],'rows':[['B','=1+2']]}]});self.assertEqual((appended['status'],appended['reason']),('refused','validation_failure'))
 def test_apply_rechecks_explicit_scalar_mode_for_forged_plans(self):
  snap=self.tool.inspect(self.source,view='region',sheet='Data',range_ref='A1:B2');target=next(x for x in snap['cells'] if x['coordinate']=='B2');base=self.tool.plan(snap,{'operations':[{'type':'set_cell_value','target_id':target['id'],'value':3,'expected_kind':'value'}]})['plan']
  forged=copy.deepcopy(base);forged['operations'][0]['value']='=1+2';raw=dict(forged);raw.pop('plan_sha256');forged['plan_sha256']=_object_sha(raw);out=self.root/'forged-formula-value.xlsx';result=self.tool.apply(self.source,forged,out);self.assertEqual((result['status'],result['reason']),('refused','validation_failure'));self.assertFalse(out.exists())
  append=self.tool.plan(snap,{'operations':[{'type':'append_rows','region_id':snap['region_id'],'copy_from_row_id':snap['rows'][1]['id'],'rows':[['B',3]]}]})['plan'];append['operations'][0]['rows'][0][1]='=1+2';raw=dict(append);raw.pop('plan_sha256');append['plan_sha256']=_object_sha(raw);out=self.root/'forged-append-formula-value.xlsx';result=self.tool.apply(self.source,append,out);self.assertEqual((result['status'],result['reason']),('refused','validation_failure'));self.assertFalse(out.exists())
 def test_create_and_plan_cardinality_budgets_refuse_before_work(self):
  cells={f'A{index}':{'value':index} for index in range(1,250002)};out=self.root/'too-many-cells.xlsx'
  with patch('xlsx_artifact_tool.api.Workbook',side_effect=AssertionError('renderer reached')):
   result=self.tool.create({'sheets':[{'name':'Data','cells':cells}]},out)
  self.assertEqual((result['status'],result['reason']),('refused','unsafe_plan'));self.assertFalse(out.exists())
  snap=self.tool.inspect(self.source,view='region',sheet='Data',range_ref='A1:B2');target=next(x for x in snap['cells'] if x['coordinate']=='B2');operation={'type':'set_cell_value','target_id':target['id'],'value':3,'expected_kind':'value'};result=self.tool.plan(snap,{'operations':[operation]*1001});self.assertEqual((result['status'],result['reason']),('refused','unsafe_plan'))
 def test_create_refuses_non_finite_layout_dimensions(self):
  for index,sheet in enumerate((
   {'name':'Data','cells':{},'column_widths':{'A':float('nan')}},
   {'name':'Data','cells':{},'column_widths':{'A':float('inf')}},
   {'name':'Data','cells':{},'row_heights':{'1':float('nan')}},
   {'name':'Data','cells':{},'row_heights':{'1':float('-inf')}},
  )):
   out=self.root/f'bad-layout-{index}.xlsx'
   with patch('xlsx_artifact_tool.api.Workbook',side_effect=AssertionError('renderer reached')):
    result=self.tool.create({'sheets':[sheet]},out)
   self.assertEqual((result['status'],result['reason']),('refused','validation_failure'));self.assertFalse(out.exists())
 def test_create_and_apply_refuse_non_xlsx_output(self):
  bad_create=self.root/'created.txt';result=self.tool.create({'sheets':[{'name':'Data','cells':{'A1':{'value':'A'}}}]},bad_create);self.assertEqual(result['status'],'refused');self.assertFalse(bad_create.exists())
  snap=self.tool.inspect(self.source,view='region',sheet='Data',range_ref='A1:B2');target=next(x for x in snap['cells'] if x['coordinate']=='B2');plan=self.tool.plan(snap,{'operations':[{'type':'set_cell_value','target_id':target['id'],'value':3,'expected_kind':'value'}]})['plan'];bad_apply=self.root/'applied.txt';result=self.tool.apply(self.source,plan,bad_apply);self.assertEqual(result['status'],'refused');self.assertFalse(bad_apply.exists())

 def test_forged_structural_plan_cannot_import_rows_outside_bound_region(self):
  source=self.root/'outside-row.xlsx';self.tool.create({'sheets':[{'name':'Data','cells':{'A1':{'value':'A'},'B1':{'value':1},'A2':{'value':'B'},'B2':{'value':2},'A3':{'value':'OUTSIDE'},'B3':{'value':999}}}]},source);snap=self.tool.inspect(source,view='region',sheet='Data',range_ref='A1:B2');outside=self.tool.inspect(source,view='region',sheet='Data',range_ref='A3:B3')['rows'][0]
  plan=self.tool.plan(snap,{'operations':[{'type':'reorder_rows','region_id':snap['region_id'],'row_ids':[snap['rows'][1]['id'],snap['rows'][0]['id']]}]})['plan'];forged=copy.deepcopy(plan)
  for row,allowed_id in zip(forged['operations'][0]['ordered_rows'],forged['operations'][0]['row_ids']):row.clear();row.update(copy.deepcopy(outside));row['id']=allowed_id
  raw=dict(forged);raw.pop('plan_sha256');forged['plan_sha256']=_object_sha(raw);out=self.root/'must-not-import-outside.xlsx';result=self.tool.apply(source,forged,out);self.assertEqual(result['status'],'refused');self.assertFalse(out.exists())
 def test_forged_structural_plan_cannot_change_bound_region_or_row_permutation(self):
  source=self.root/'structural-forge.xlsx';self.tool.create({'sheets':[{'name':'Data','cells':{'A1':{'value':'Head'},'B1':{'value':0},'A2':{'value':'KEEP'},'B2':{'value':99}}}]},source);snap=self.tool.inspect(source,view='region',sheet='Data',range_ref='A1:B2')
  append=self.tool.plan(snap,{'operations':[{'type':'append_rows','region_id':snap['region_id'],'copy_from_row_id':snap['rows'][1]['id'],'rows':[['NEW',1]]}]})['plan'];append['operations'][0]['range']='A1:B1';raw=dict(append);raw.pop('plan_sha256');append['plan_sha256']=_object_sha(raw);out=self.root/'forged-append.xlsx';result=self.tool.apply(source,append,out);self.assertEqual(result['status'],'refused');self.assertFalse(out.exists())
  reorder=self.tool.plan(snap,{'operations':[{'type':'reorder_rows','region_id':snap['region_id'],'row_ids':[snap['rows'][1]['id'],snap['rows'][0]['id']]}]})['plan'];reorder['operations'][0]['ordered_rows']=[reorder['operations'][0]['source_rows'][0]]*2;reorder['operations'][0]['row_ids']=[snap['rows'][0]['id']]*2;raw=dict(reorder);raw.pop('plan_sha256');reorder['plan_sha256']=_object_sha(raw);out=self.root/'forged-reorder.xlsx';result=self.tool.apply(source,reorder,out);self.assertEqual(result['status'],'refused');self.assertFalse(out.exists())
 def test_append_candidate_postcondition_blocks_publication(self):
  snap=self.tool.inspect(self.source,view='region',sheet='Data',range_ref='A1:B2');plan=self.tool.plan(snap,{'operations':[{'type':'append_rows','region_id':snap['region_id'],'copy_from_row_id':snap['rows'][1]['id'],'rows':[['B',3]]}]})['plan'];out=self.root/'no-append.xlsx';real=load_workbook
  def corrupt(path,*args,**kwargs):
   wb=real(path,*args,**kwargs)
   if Path(path)!=self.source and wb['Data'].max_row>=3:wb['Data']['B3']=999
   return wb
  with patch('xlsx_artifact_tool.api.load_workbook',side_effect=corrupt):result=self.tool.apply(self.source,plan,out)
  self.assertEqual(result['status'],'refused');self.assertEqual(result['reason'],'validation_failure');self.assertFalse(out.exists())

 def test_create_candidate_semantics_are_validated_before_publication(self):
  out=self.root/'wrong-create.xlsx';real_save=__import__('openpyxl',fromlist=['Workbook']).Workbook.save
  def save_wrong(workbook,path):
   from openpyxl import Workbook
   wrong=Workbook();wrong.active.title='Wrong';wrong.active['A1']='WRONG';return real_save(wrong,path)
  with patch('openpyxl.workbook.workbook.Workbook.save',new=save_wrong):result=self.tool.create({'sheets':[{'name':'Expected','cells':{'A1':{'value':'RIGHT'}}}]},out)
  self.assertEqual(result['status'],'refused');self.assertFalse(out.exists())
 def test_create_candidate_style_and_layout_are_validated_before_publication(self):
  out=self.root/'wrong-create-style.xlsx'
  with patch('xlsx_artifact_tool.api._apply_style',return_value=None):result=self.tool.create({'sheets':[{'name':'Expected','cells':{'A1':{'value':'RIGHT','style':'header'}},'column_widths':{'A':24},'row_heights':{'1':28}}]},out)
  self.assertEqual(result['status'],'refused');self.assertFalse(out.exists())
 def test_create_candidate_must_be_openable_before_publication(self):
  out=self.root/'never-created.xlsx';real_admit=__import__('xlsx_artifact_tool.api',fromlist=['_admit'])._admit
  def corrupt_after_admission(path):
   real_admit(path)
   Path(path).write_bytes(b'not an xlsx')
  with patch('xlsx_artifact_tool.api._admit',side_effect=corrupt_after_admission):result=self.tool.create({'sheets':[{'name':'Data','cells':{'A1':{'value':'A'}}}]},out)
  self.assertEqual((result['status'],result['reason']),('refused','validation_failure'));self.assertFalse(out.exists())
 def test_multiple_structural_edits_on_one_region_are_refused(self):
  snap=self.tool.inspect(self.source,view='region',sheet='Data',range_ref='A1:B2');template=snap['rows'][1]
  append_op=self.tool.plan(snap,{'operations':[{'type':'append_rows','region_id':snap['region_id'],'copy_from_row_id':template['id'],'rows':[['B',3]]}]})['plan']['operations'][0]
  reorder_op=self.tool.plan(snap,{'operations':[{'type':'reorder_rows','region_id':snap['region_id'],'row_ids':[snap['rows'][1]['id'],snap['rows'][0]['id']]}]})['plan']['operations'][0]
  requested={'operations':[{'type':'append_rows','region_id':snap['region_id'],'copy_from_row_id':template['id'],'rows':[['B',3]]},{'type':'reorder_rows','region_id':snap['region_id'],'row_ids':[snap['rows'][1]['id'],snap['rows'][0]['id']]}]}
  result=self.tool.plan(snap,requested);self.assertEqual((result['status'],result['reason']),('refused','conflict'))
  forged={'schema':1,'source_sha256':snap['source_sha256'],'snapshot_sha256':snap['snapshot_sha256'],'operations':[append_op,reorder_op]};forged['plan_sha256']=_object_sha(forged);out=self.root/'forged-structural-conflict.xlsx';result=self.tool.apply(self.source,forged,out);self.assertEqual((result['status'],result['reason']),('refused','conflict'));self.assertFalse(out.exists())
 def test_normal_style_clears_managed_header_font_and_fill(self):
  source=self.root/'normal-style.xlsx';self.tool.create({'sheets':[{'name':'Data','cells':{'A1':{'value':'Heading','style':'header'}}}]},source);snap=self.tool.inspect(source,view='region',sheet='Data',range_ref='A1:A1');target=snap['cells'][0];plan=self.tool.plan(snap,{'operations':[{'type':'set_cell_style','target_id':target['id'],'style':'normal'}]})['plan'];out=self.root/'normal-style-out.xlsx';result=self.tool.apply(source,plan,out);self.assertEqual(result['status'],'ok');cell=load_workbook(out)['Data']['A1'];self.assertFalse(cell.font.bold);self.assertIsNone(cell.fill.fill_type);self.assertEqual(cell.number_format,'General')
 def test_style_candidate_oracle_detects_noop_and_validate_allows_styles_part(self):
  snap=self.tool.inspect(self.source,view='region',sheet='Data',range_ref='A1:B2');target=next(x for x in snap['cells'] if x['coordinate']=='B2');plan=self.tool.plan(snap,{'operations':[{'type':'set_cell_style','target_id':target['id'],'style':'header'}]})['plan'];out=self.root/'no-style.xlsx'
  with patch('xlsx_artifact_tool.api._apply_style',return_value=None):result=self.tool.apply(self.source,plan,out)
  self.assertEqual(result['status'],'refused');self.assertFalse(out.exists())
  plan=self.tool.plan(snap,{'operations':[{'type':'set_cell_style','target_id':target['id'],'style':'currency'}]})['plan'];out=self.root/'currency.xlsx';self.assertEqual(self.tool.apply(self.source,plan,out)['status'],'ok');report=self.tool.validate(out,before=self.source);self.assertEqual(report['status'],'valid');self.assertEqual(report['unexpected_changed_members'],[])
 def test_candidate_semantic_failure_happens_before_publication(self):
  snap=self.tool.inspect(self.source,view='region',sheet='Data',range_ref='A1:B2');target=next(x for x in snap['cells'] if x['coordinate']=='B2');plan=self.tool.plan(snap,{'operations':[{'type':'set_cell_value','target_id':target['id'],'value':3,'expected_kind':'value'}]})['plan'];out=self.root/'never.xlsx'
  real=load_workbook
  def corrupt_candidate(path,*args,**kwargs):
   wb=real(path,*args,**kwargs)
   if '.candidate.' in Path(path).name:wb['Data']['B2']=999
   return wb
  import os
  real_replace=os.replace;publication_attempts=[]
  def guard_publication(src,dst):
   if Path(dst)==out:publication_attempts.append((src,dst));raise AssertionError('published')
   return real_replace(src,dst)
  with patch('xlsx_artifact_tool.api.load_workbook',side_effect=corrupt_candidate),patch('xlsx_artifact_tool.api.os.replace',side_effect=guard_publication):
   result=self.tool.apply(self.source,plan,out)
  self.assertEqual(result['status'],'refused');self.assertEqual(result['reason'],'validation_failure');self.assertFalse(out.exists());self.assertEqual(publication_attempts,[])

if __name__=='__main__':unittest.main()
