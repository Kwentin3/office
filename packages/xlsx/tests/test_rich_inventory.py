from __future__ import annotations

import copy
import json
import tempfile
import unittest
import zipfile
from pathlib import Path

from lxml import etree
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from xlsx_artifact_tool import XlsxArtifactTool
from xlsx_artifact_tool.inventory import BLOCKER_FEATURES, WARNING_FEATURES

XLSX_FEATURES = {
    "external_links",
    "external_relationships",
    "charts",
    "pivots",
    "tables",
    "comments",
    "data_validations",
    "conditional_formatting",
    "defined_names",
    "merged_ranges",
    "formula_cells",
    "shared_formulas",
    "array_formulas",
    "drawings",
    "connections",
    "embedded_packages",
    "ole_objects",
    "activex",
    "workbook_protection",
    "sheet_protection",
    "macros",
    "signatures",
}


class RichInventoryTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.tool = XlsxArtifactTool(self.root / "work")
        self.source = self.root / "source.xlsx"
        workbook = Workbook()
        workbook.active["A1"] = "Normal"
        workbook.save(self.source)

    def tearDown(self) -> None:
        self.temp.cleanup()

    def add_member(self, name: str, payload: bytes = b"x") -> None:
        rewritten = self.root / "rewritten.xlsx"
        with zipfile.ZipFile(self.source) as source, zipfile.ZipFile(rewritten, "w") as output:
            for info in source.infolist():
                output.writestr(copy.copy(info), source.read(info.filename))
            output.writestr(name, payload)
        rewritten.replace(self.source)

    def add_ole_node(self) -> None:
        rewritten = self.root / "rewritten.xlsx"
        namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        with zipfile.ZipFile(self.source) as source, zipfile.ZipFile(rewritten, "w") as output:
            for info in source.infolist():
                payload = source.read(info.filename)
                if info.filename == "xl/worksheets/sheet1.xml":
                    root = etree.fromstring(payload)
                    container = etree.SubElement(root, f"{{{namespace}}}oleObjects")
                    node = etree.SubElement(container, f"{{{namespace}}}oleObject")
                    node.set("progId", "Package")
                    payload = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)
                output.writestr(copy.copy(info), payload)
        rewritten.replace(self.source)

    def test_inventory_reports_safe_baseline(self) -> None:
        result = self.tool.inspect(self.source, view="inventory")
        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["schema_version"], 2)
        self.assertEqual(set(result["features"]), XLSX_FEATURES)
        self.assertEqual(result["mutation_policy"]["decision"], "safe")
        self.assertEqual(result["features"]["external_links"], 0)
        self.assertEqual(result["features"]["charts"], 0)
        self.assertEqual(result["findings"], [])
        self.assertFalse(result["findings_truncated"])

    def test_inventory_v2_findings_are_deterministic_and_location_aware(self) -> None:
        workbook = Workbook()
        data = workbook.active
        data.title = "Data"
        data["A1"] = "amount"
        data["A2"] = 10
        data["A2"].comment = Comment("review", "tester")
        validation = DataValidation(type="whole", operator="greaterThan", formula1="0")
        validation.add("A2:A9")
        data.add_data_validation(validation)
        data.conditional_formatting.add("A2:A9", CellIsRule(operator="greaterThan", formula=["5"]))
        data.merge_cells("C1:D1")
        other = workbook.create_sheet("Other")
        other["A1"] = "untouched"
        workbook.create_named_range("InputRange", data, "A2:A9")
        workbook.save(self.source)

        first = self.tool.inspect(self.source, view="inventory")
        second = self.tool.inspect(self.source, view="inventory")
        self.assertEqual(first["findings"], second["findings"])
        self.assertFalse(first["findings_truncated"])
        by_feature = {}
        for finding in first["findings"]:
            by_feature.setdefault(finding["feature"], []).append(finding)
            self.assertFalse(finding["part"].startswith("/"))
            self.assertNotIn("..", Path(finding["part"]).parts)
        self.assertEqual(by_feature["comments"][0]["scope"], "worksheet")
        self.assertEqual(by_feature["comments"][0]["sheet"], "Data")
        self.assertEqual(by_feature["comments"][0]["range"], "A2")
        self.assertTrue(by_feature["comments"][0]["part"].startswith("xl/comments"))
        self.assertEqual(by_feature["data_validations"][0]["range"], "A2:A9")
        self.assertEqual(by_feature["conditional_formatting"][0]["range"], "A2:A9")
        self.assertEqual(by_feature["merged_ranges"][0]["range"], "C1:D1")
        self.assertEqual(by_feature["defined_names"][0]["sheet"], "Data")
        self.assertEqual(by_feature["defined_names"][0]["range"], "A2:A9")

    def test_inventory_v2_findings_are_bounded_and_report_truncation(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        for row in range(1, 1102):
            sheet.cell(row, 1, f"={row}")
        workbook.save(self.source)
        result = self.tool.inspect(self.source, view="inventory")
        self.assertEqual(result["features"]["formula_cells"], 1101)
        self.assertEqual(len(result["findings"]), 1000)
        self.assertTrue(result["findings_truncated"])
        self.assertEqual(
            result["findings"],
            sorted(
                result["findings"],
                key=lambda item: (
                    item["feature"], item["scope"], item["part"], item.get("sheet", ""), item.get("range", "")
                ),
            ),
        )

    def test_inventory_schema_matches_policy_partitions(self) -> None:
        schema = json.loads(
            (Path(__file__).parents[1] / "xlsx_artifact_tool/resources/inventory.schema.json").read_text()
        )
        self.assertEqual(schema["properties"]["schema_version"]["const"], 2)
        self.assertEqual(schema["properties"]["findings"]["maxItems"], 1000)
        self.assertIn("findings_truncated", schema["required"])
        policy = schema["properties"]["mutation_policy"]
        self.assertEqual(set(policy["properties"]["blockers"]["items"]["enum"]), set(BLOCKER_FEATURES))
        self.assertEqual(set(policy["properties"]["warnings"]["items"]["enum"]), set(WARNING_FEATURES))
        constraints = {
            rule["if"]["properties"]["decision"]["const"]: rule["then"]["properties"] for rule in policy["allOf"]
        }
        self.assertEqual(constraints["safe"]["blockers"]["maxItems"], 0)
        self.assertEqual(constraints["safe"]["warnings"]["maxItems"], 0)
        self.assertEqual(constraints["safe_with_warnings"]["warnings"]["minItems"], 1)
        self.assertEqual(constraints["refuse_mutation"]["blockers"]["minItems"], 1)

    def test_malformed_package_xml_is_typed_for_inventory_and_apply(self) -> None:
        snapshot = self.tool.inspect(self.source, view="region", sheet="Sheet", range_ref="A1:A1")
        target = snapshot["cells"][0]
        planned = self.tool.plan(
            snapshot,
            {
                "operations": [
                    {"type": "set_cell_value", "target_id": target["id"], "value": "Changed", "expected_kind": "value"}
                ]
            },
        )
        rewritten = self.root / "malformed.xlsx"
        with zipfile.ZipFile(self.source) as source, zipfile.ZipFile(rewritten, "w") as output:
            for info in source.infolist():
                payload = b"<broken" if info.filename == "docProps/core.xml" else source.read(info.filename)
                output.writestr(copy.copy(info), payload)
        rewritten.replace(self.source)
        inventory = self.tool.inspect(self.source, view="inventory")
        self.assertEqual((inventory["status"], inventory["reason"]), ("refused", "validation_failure"))
        result = self.tool.apply(self.source, planned["plan"], self.root / "out.xlsx")
        self.assertEqual((result["status"], result["reason"]), ("refused", "validation_failure"))

    def test_ole_objects_are_inventory_blocker_and_apply_refuses(self) -> None:
        snapshot = self.tool.inspect(self.source, view="region", sheet="Sheet", range_ref="A1:A1")
        target = snapshot["cells"][0]
        planned = self.tool.plan(
            snapshot,
            {
                "operations": [
                    {"type": "set_cell_value", "target_id": target["id"], "value": "Changed", "expected_kind": "value"}
                ]
            },
        )
        self.add_ole_node()
        inventory = self.tool.inspect(self.source, view="inventory")
        self.assertEqual(inventory["features"]["ole_objects"], 1)
        self.assertEqual(inventory["mutation_policy"]["decision"], "refuse_mutation")
        result = self.tool.apply(self.source, planned["plan"], self.root / "out.xlsx")
        self.assertEqual(result["status"], "refused")
        self.assertEqual(result["reason"], "unsupported_capability")

    def test_embedded_package_is_warning_not_ole_blocker(self) -> None:
        self.add_member("xl/embeddings/chart-workbook.xlsx")
        inventory = self.tool.inspect(self.source, view="inventory")
        self.assertEqual(inventory["features"]["embedded_packages"], 1)
        self.assertEqual(inventory["features"]["ole_objects"], 0)
        self.assertEqual(inventory["mutation_policy"]["decision"], "safe_with_warnings")

    def test_sheet_protection_is_blocker(self) -> None:
        workbook = Workbook()
        workbook.active["A1"] = "Normal"
        workbook.active.protection.sheet = True
        workbook.save(self.source)
        inventory = self.tool.inspect(self.source, view="inventory")
        self.assertEqual(inventory["features"]["sheet_protection"], 1)
        self.assertEqual(inventory["mutation_policy"]["decision"], "refuse_mutation")


if __name__ == "__main__":
    unittest.main()
