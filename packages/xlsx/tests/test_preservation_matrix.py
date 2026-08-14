from __future__ import annotations

import copy
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from xlsx_artifact_tool import XlsxArtifactTool


class PreservationMatrixTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.tool = XlsxArtifactTool(self.root / "work")

    def tearDown(self) -> None:
        self.temp.cleanup()

    def workbook_with_comment(self, comment_sheet: str = "Data") -> Path:
        source = self.root / f"comment-{comment_sheet}.xlsx"
        workbook = Workbook()
        data = workbook.active
        data.title = "Data"
        data.append(["Item", "Amount"])
        data.append(["A", 2])
        other = workbook.create_sheet("Other")
        other["A1"] = "untouched"
        workbook[comment_sheet]["A1"].comment = Comment("preserve", "tester")
        workbook.save(source)
        return source

    @staticmethod
    def members(path: Path) -> dict[str, bytes]:
        with zipfile.ZipFile(path) as archive:
            return {info.filename: archive.read(info.filename) for info in archive.infolist()}

    def plan_cell(self, source: Path, operation: dict) -> dict:
        snapshot = self.tool.inspect(source, view="region", sheet="Data", range_ref="A1:B2")
        target = next(cell for cell in snapshot["cells"] if cell["coordinate"] == "B2")
        request = dict(operation)
        request["target_id"] = target["id"]
        if request["type"] != "set_cell_style":
            request.setdefault("expected_kind", target["kind"])
        planned = self.tool.plan(snapshot, {"operations": [request]})
        self.assertEqual(planned["status"], "ok")
        self.assertEqual(planned["plan"]["schema"], 1)
        return planned["plan"]

    def assert_exact_allowlist(self, before: dict[str, bytes], after: dict[str, bytes], allowlist: set[str]) -> None:
        self.assertEqual(set(after), set(before))
        changed = {name for name in before if before[name] != after[name]}
        self.assertTrue(changed)
        self.assertLessEqual(changed, allowlist)
        for name in set(before) - allowlist:
            self.assertEqual(after[name], before[name], name)

    def test_clear_cell_uses_package_preserving_path_with_comments(self) -> None:
        source = self.workbook_with_comment("Data")
        before = self.members(source)
        plan = self.plan_cell(source, {"type": "clear_cell"})
        output = self.root / "clear.xlsx"
        result = self.tool.apply(source, plan, output)
        self.assertEqual(result["status"], "ok")
        self.assertIsNone(load_workbook(output)["Data"]["B2"].value)
        self.assertEqual(load_workbook(output)["Data"]["A1"].comment.text, "preserve")
        self.assert_exact_allowlist(before, self.members(output), {"xl/worksheets/sheet1.xml"})

    def test_style_cell_preserves_rich_package_outside_style_allowlist(self) -> None:
        source = self.workbook_with_comment("Data")
        before = self.members(source)
        plan = self.plan_cell(source, {"type": "set_cell_style", "style": "currency"})
        output = self.root / "style.xlsx"
        result = self.tool.apply(source, plan, output)
        self.assertEqual(result["status"], "ok")
        after = load_workbook(output)
        self.assertEqual(after["Data"]["B2"].number_format, "#,##0.00")
        self.assertEqual(after["Data"]["A1"].comment.text, "preserve")
        self.assert_exact_allowlist(before, self.members(output), {"xl/worksheets/sheet1.xml", "xl/styles.xml"})

    def test_structural_policy_uses_exact_affected_worksheet(self) -> None:
        source = self.workbook_with_comment("Other")
        before = self.members(source)
        snapshot = self.tool.inspect(source, view="region", sheet="Data", range_ref="A1:B2")
        request = {
            "operations": [
                {
                    "type": "append_rows",
                    "region_id": snapshot["region_id"],
                    "copy_from_row_id": snapshot["rows"][1]["id"],
                    "rows": [["B", 3]],
                }
            ]
        }
        plan = self.tool.plan(snapshot, request)["plan"]
        self.assertEqual(plan["schema"], 1)
        output = self.root / "append-other-comment.xlsx"
        result = self.tool.apply(source, plan, output)
        self.assertEqual(result["status"], "ok")
        after = load_workbook(output)
        self.assertEqual(after["Data"]["B3"].value, 3)
        self.assertEqual(after["Other"]["A1"].comment.text, "preserve")
        self.assert_exact_allowlist(before, self.members(output), {"xl/worksheets/sheet1.xml"})

    def test_structural_policy_refuses_feature_on_affected_worksheet_without_output(self) -> None:
        source = self.workbook_with_comment("Data")
        snapshot = self.tool.inspect(source, view="region", sheet="Data", range_ref="A1:B2")
        plan = self.tool.plan(
            snapshot,
            {
                "operations": [
                    {
                        "type": "reorder_rows",
                        "region_id": snapshot["region_id"],
                        "row_ids": [snapshot["rows"][1]["id"], snapshot["rows"][0]["id"]],
                    }
                ]
            },
        )["plan"]
        output = self.root / "refused.xlsx"
        result = self.tool.apply(source, plan, output)
        self.assertEqual((result["status"], result["reason"]), ("refused", "unsupported_capability"))
        self.assertFalse(output.exists())

    def test_unsafe_external_relationship_is_global_for_package_safe_edits(self) -> None:
        source = self.workbook_with_comment("Other")
        rewritten = self.root / "external.xlsx"
        with zipfile.ZipFile(source) as archive, zipfile.ZipFile(rewritten, "w") as output:
            for info in archive.infolist():
                payload = archive.read(info.filename)
                if info.filename == "xl/_rels/workbook.xml.rels":
                    root = __import__("lxml.etree", fromlist=["etree"]).fromstring(payload)
                    relationship = __import__("lxml.etree", fromlist=["etree"]).SubElement(
                        root, "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
                    )
                    relationship.set("Id", "rIdUnsafe")
                    relationship.set("Type", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink")
                    relationship.set("Target", "https://example.invalid")
                    relationship.set("TargetMode", "External")
                    payload = __import__("lxml.etree", fromlist=["etree"]).tostring(root, xml_declaration=True, encoding="UTF-8")
                output.writestr(copy.copy(info), payload)
        plan = self.plan_cell(rewritten, {"type": "set_cell_value", "value": 9})
        output = self.root / "external-refused.xlsx"
        result = self.tool.apply(rewritten, plan, output)
        self.assertEqual((result["status"], result["reason"]), ("refused", "unsupported_capability"))
        self.assertFalse(output.exists())


if __name__ == "__main__":
    unittest.main()
