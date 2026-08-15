"""Bounded, script-free HTML previews for validated XLSX create models."""

from __future__ import annotations

import ctypes
import hashlib
import html
import json
import os
import shutil
import tempfile
from pathlib import Path
from typing import Any

from openpyxl.utils.cell import get_column_letter, range_boundaries

from .api import validate_create_model
from .review_contract import validate_review_packet


class PreviewError(ValueError):
    """A deterministic XLSX preview refusal."""


def _exchange_directories(source: Path, target: Path) -> None:
    """Atomically exchange two existing directories or fail closed."""
    libc = ctypes.CDLL(None, use_errno=True)
    renameat2 = getattr(libc, "renameat2", None)
    if renameat2 is None:
        raise PreviewError("atomic directory exchange is unavailable")
    renameat2.argtypes = [ctypes.c_int, ctypes.c_char_p, ctypes.c_int, ctypes.c_char_p, ctypes.c_uint]
    renameat2.restype = ctypes.c_int
    if renameat2(-100, os.fsencode(source), -100, os.fsencode(target), 2) != 0:
        error = ctypes.get_errno()
        raise OSError(error, os.strerror(error))


def _cleanup_old_generation(temporary: Path, backup: Path) -> None:
    """Best-effort cleanup after the directory exchange commit point."""
    try:
        os.replace(temporary, backup)
    except Exception:
        try:
            shutil.rmtree(temporary)
        except Exception:
            pass
        return
    try:
        shutil.rmtree(backup)
    except Exception:
        pass


_MAX_PREVIEW_ROWS = 100
_MAX_PREVIEW_COLUMNS = 40
_MAX_PREVIEW_SHEETS = 20


def _digest_json(value: Any) -> str:
    encoded = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _file_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _escape(value: Any) -> str:
    return html.escape("" if value is None else str(value), quote=True)


def _sheet_html(sheet: dict[str, Any]) -> tuple[str, dict[str, Any] | None]:
    addressed: dict[tuple[int, int], tuple[str, dict[str, Any]]] = {}
    maximum_row = maximum_column = 1
    for source_coordinate, payload in sheet["cells"].items():
        min_column, min_row, _, _ = range_boundaries(source_coordinate)
        coordinate = f"{get_column_letter(min_column)}{min_row}"
        addressed[(min_row, min_column)] = (coordinate, payload)
        maximum_row = max(maximum_row, min_row)
        maximum_column = max(maximum_column, min_column)

    rows_shown = min(maximum_row, _MAX_PREVIEW_ROWS)
    columns_shown = min(maximum_column, _MAX_PREVIEW_COLUMNS)
    omitted_cells = sum(
        1 for row, column in addressed if row > _MAX_PREVIEW_ROWS or column > _MAX_PREVIEW_COLUMNS
    )

    sheet_name = _escape(sheet["name"])
    out = [
        "<!doctype html>",
        '<html lang="en"><head><meta charset="utf-8">',
        f"<title>{sheet_name}</title>",
        "<style>body{font-family:system-ui,sans-serif;margin:24px;color:#172033}"
        "table{border-collapse:collapse}th,td{border:1px solid #cbd5e1;padding:6px;min-width:72px}"
        "th{background:#f1f5f9}td.formula{color:#4338ca}</style></head><body>",
        f'<h1 data-sheet="{sheet_name}">{sheet_name}</h1>',
        f'<table data-sheet="{sheet_name}"><thead><tr><th></th>',
    ]
    out.extend(f"<th>{get_column_letter(column)}</th>" for column in range(1, columns_shown + 1))
    out.append("</tr></thead><tbody>")
    for row in range(1, rows_shown + 1):
        out.append(f"<tr><th>{row}</th>")
        for column in range(1, columns_shown + 1):
            coordinate = f"{get_column_letter(column)}{row}"
            payload = addressed.get((row, column))
            value = "" if payload is None else payload[1].get("formula", payload[1].get("value"))
            kind = "formula" if payload is not None and "formula" in payload[1] else "value"
            out.append(
                f'<td class="{kind}" data-sheet="{sheet_name}" data-cell="{coordinate}">{_escape(value)}</td>'
            )
        out.append("</tr>")
    out.append("</tbody></table></body></html>\n")
    diagnostic = None
    if omitted_cells:
        diagnostic = {
            "sheet": sheet["name"],
            "rows_shown": rows_shown,
            "columns_shown": columns_shown,
            "omitted_cells": omitted_cells,
        }
    return "".join(out), diagnostic


def render_xlsx_preview(model: dict[str, Any], output_dir: str | Path) -> dict[str, Any]:
    """Validate an XLSX model and atomically publish chat-only review evidence."""
    try:
        validated = validate_create_model(model)
    except ValueError as exc:
        reason = str(exc) if str(exc) in {"validation_failure", "unsafe_plan"} else "validation_failure"
        return {"status": "refused", "reason": reason, "details": "invalid XLSX create model"}
    revision = _digest_json(validated)
    workbook_id = validated.get("workbook_id", f"workbook_{revision[:16]}")
    raw_output = Path(output_dir).expanduser().absolute()
    raw_backup = raw_output.with_name(f".{raw_output.name}.old")
    if raw_output.is_symlink() or raw_backup.is_symlink():
        raise PreviewError("preview output and backup must not be symlinks")
    output = raw_output.resolve(strict=False)
    backup = raw_backup.resolve(strict=False)
    if output.exists() and not output.is_dir():
        raise PreviewError("existing preview output must be a directory")
    if backup.exists() and not backup.is_dir():
        raise PreviewError("preview backup path must be a directory")
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = Path(tempfile.mkdtemp(prefix=f".{output.name}.", dir=output.parent))
    try:
        review_sheets: list[dict[str, Any]] = []
        truncated_sheets: list[dict[str, Any]] = []
        omitted_sheets = [
            {"sheet": sheet["name"], "reason": "very_hidden" if sheet.get("state") == "veryHidden" else "hidden"}
            for sheet in validated["sheets"]
            if sheet.get("state", "visible") != "visible"
        ]
        visible_sheets = [sheet for sheet in validated["sheets"] if sheet.get("state", "visible") == "visible"]
        omitted_sheets.extend(
            {"sheet": sheet["name"], "reason": "artifact_limit"}
            for sheet in visible_sheets[_MAX_PREVIEW_SHEETS:]
        )
        for number, sheet in enumerate(visible_sheets[:_MAX_PREVIEW_SHEETS], start=1):
            filename = f"sheet-{number:02d}.html"
            path = temporary / filename
            rendered, truncation = _sheet_html(sheet)
            path.write_text(rendered, encoding="utf-8")
            if truncation is not None:
                truncated_sheets.append(truncation)
            review_sheets.append(
                {
                    "sheet": sheet["name"],
                    "number": number,
                    "html_file": filename,
                    "html_sha256": _file_sha256(path),
                }
            )
        review = validate_review_packet(
            {
                "contract_version": "1.0",
                "kind": "xlsx_chat_review",
                "interaction": "chat_only",
                "workbook_id": workbook_id,
                "revision": revision,
                "fidelity": "structural_preview_not_excel_render",
                "limitations": [
                    "Formula text is displayed but never evaluated",
                    "Formatting and layout are structural approximations, not an Excel render",
                ],
                "diagnostics": {"truncated_sheets": truncated_sheets, "omitted_sheets": omitted_sheets},
                "sheets": review_sheets,
            }
        )
        (temporary / "review.json").write_text(
            json.dumps(review, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        if backup.exists():
            shutil.rmtree(backup)
        if output.exists():
            _exchange_directories(temporary, output)
            _cleanup_old_generation(temporary, backup)
        else:
            os.replace(temporary, output)
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise

    return {
        "status": "previewed",
        "workbook_id": workbook_id,
        "interaction": "chat_only",
        "revision": revision,
        "review_contract": str((output / "review.json").resolve()),
        "display_artifacts": [str((output / sheet["html_file"]).resolve()) for sheet in review_sheets],
        "output": str(output),
    }
