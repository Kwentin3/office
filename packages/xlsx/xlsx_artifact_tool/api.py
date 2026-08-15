from __future__ import annotations

import copy
import hashlib
import json
import math
import os
import re
import stat
import tempfile
import zipfile
from pathlib import Path, PurePosixPath
from typing import Any

from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import get_column_letter, range_boundaries

from .inventory import inspect_inventory
from .preservation import PACKAGE_CELL_OPERATIONS, admit_operations
from .template import (
 has_marker,
 package_has_marker,
 token_names,
 unsupported_scope_has_marker,
 validate_values,
 well_formed,
)
from .template import (
 render as render_tokens,
)

_STYLES={
 'normal':{'font':Font(),'fill':PatternFill(fill_type=None,fgColor='00000000'),'number_format':'General'},
 'header':{'font':Font(bold=True,color='FFFFFF'),'fill':PatternFill('solid',fgColor='4472C4')},
 'currency':{'number_format':'#,##0.00'},'percent':{'number_format':'0.00%'},'date':{'number_format':'yyyy-mm-dd'},
 'integer':{'number_format':'0'},'text':{'number_format':'@'},
}
_FORMULA_EXTERNAL=re.compile(r"\[[^\]]+\]")
_FORMULA_DDE=re.compile(r"(?i)^=\s*(?:[a-z_][\w.]*\s*\||ddeauto\s*\()")
_WORKBOOK_ID=re.compile(r"^[A-Za-z][A-Za-z0-9_.-]{0,63}$")
_CELL_REF=re.compile(r"^[A-Z]{1,3}[1-9][0-9]{0,6}$")
_COLUMN_REF=re.compile(r"^[A-Z]{1,3}$")
_ROW_REF=re.compile(r"^[1-9][0-9]{0,6}$")

def _single_cell_bounds(value):
 if not isinstance(value,str) or not _CELL_REF.fullmatch(value):raise ValueError('validation_failure')
 try:min_col,min_row,max_col,max_row=range_boundaries(value)
 except Exception as exc:raise ValueError('validation_failure') from exc
 if None in (min_col,min_row,max_col,max_row) or min_col!=max_col or min_row!=max_row or max_col>16384 or max_row>1048576 or f'{get_column_letter(min_col)}{min_row}'!=value:raise ValueError('validation_failure')
 return min_col,min_row,max_col,max_row

def _cell_range_bounds(value,*,allow_single):
 if not isinstance(value,str):raise ValueError('validation_failure')
 parts=value.split(':')
 if len(parts) not in ({1,2} if allow_single else {2}):raise ValueError('validation_failure')
 first=_single_cell_bounds(parts[0]);last=_single_cell_bounds(parts[-1])
 min_col,min_row=first[0],first[1];max_col,max_row=last[2],last[3]
 if min_col>max_col or min_row>max_row:raise ValueError('validation_failure')
 canonical=f'{get_column_letter(min_col)}{min_row}'
 if len(parts)==2:canonical+=f':{get_column_letter(max_col)}{max_row}'
 if canonical!=value:raise ValueError('validation_failure')
 return min_col,min_row,max_col,max_row

def _parse_xml(payload:bytes):
 if b'<!DOCTYPE' in payload or b'<!ENTITY' in payload:raise ValueError('validation_failure')
 root=etree.fromstring(payload,parser=etree.XMLParser(resolve_entities=False,no_network=True,huge_tree=False));info=root.getroottree().docinfo
 if info.doctype or info.internalDTD is not None or info.externalDTD is not None:raise ValueError('validation_failure')
 return root
def _sha(path:Path):return hashlib.sha256(path.read_bytes()).hexdigest()
def _object_sha(value):return hashlib.sha256(json.dumps(value,sort_keys=True,separators=(',',':'),ensure_ascii=False).encode()).hexdigest()
_REFUSAL_REASONS={'unsupported_capability','ambiguous_target','stale_snapshot','validation_failure','unsafe_plan','conflict'}
def _typed_reason(exc):
 value=str(exc)
 return value if value in _REFUSAL_REASONS else 'validation_failure'
def _refusal(reason,details=''):return {'status':'refused','reason':reason if reason in _REFUSAL_REASONS else 'validation_failure','details':details}
def _id(source,sheet,coordinate,kind,value):return 'tx_'+hashlib.sha256(f'{source}|{sheet}|{coordinate}|{kind}|{value!r}'.encode()).hexdigest()[:24]
def _safe_formula(value):return isinstance(value,str) and value.startswith('=') and len(value)<=32767 and not _FORMULA_EXTERNAL.search(value) and not _FORMULA_DDE.search(value)
def _scalar(value):return value is None or isinstance(value,(str,bool,int)) or isinstance(value,float) and math.isfinite(value)
def _explicit_scalar(value):return _scalar(value) and (not isinstance(value,str) or not value.startswith('='))
def _row_id(source,sheet,row_number,cells):
 payload=[(item['coordinate'],item['kind'],item.get('value',item.get('formula')),item['style_id']) for item in cells]
 return 'tx_'+hashlib.sha256(f'{source}|{sheet}|row|{row_number}|{payload!r}'.encode()).hexdigest()[:24]
def _region_id(source:str,sheet:str,range_ref:str):return 'tx_'+hashlib.sha256(f'{source}|{sheet}|region|{range_ref}'.encode()).hexdigest()[:24]
def _admit(path:Path):
 try:
  if path.suffix.lower()!='.xlsx':raise ValueError
  with zipfile.ZipFile(path) as z:
   infos=z.infolist();names=[x.filename for x in infos];total=sum(x.file_size for x in infos)
   required={'[Content_Types].xml','_rels/.rels','xl/workbook.xml','xl/_rels/workbook.xml.rels'}
   if not required<=set(names):raise ValueError
   modes=[(x.external_attr>>16)&0xFFFF for x in infos]
   unsafe_type=any(not (info.is_dir() or stat.S_IFMT(mode) in (0,stat.S_IFREG)) for info,mode in zip(infos,modes))
   unsafe_size=len(infos)>10000 or total>64*1024*1024 or any(x.file_size>16*1024*1024 or x.file_size>1024*1024 and x.file_size/max(x.compress_size,1)>200 for x in infos)
   if len(names)!=len(set(names)) or any(n.startswith('/') or '\\' in n or '..' in PurePosixPath(n).parts for n in names) or unsafe_type or unsafe_size:raise ValueError
   for name in names:
    if name.endswith(('.xml','.rels')):_parse_xml(z.read(name))
 except etree.XMLSyntaxError as exc:raise ValueError('validation_failure') from exc
 except Exception as exc:raise ValueError('unsafe_package') from exc

def _candidate_openable(path:Path):
 try:
  _admit(path);workbook=load_workbook(path,data_only=False);workbook.close()
 except Exception as exc:raise ValueError('validation_failure') from exc

def _sheet_member(path:Path,sheet_name:str):
 ns={'m':'http://schemas.openxmlformats.org/spreadsheetml/2006/main','r':'http://schemas.openxmlformats.org/officeDocument/2006/relationships','p':'http://schemas.openxmlformats.org/package/2006/relationships'}
 with zipfile.ZipFile(path) as archive:
  workbook=etree.fromstring(archive.read('xl/workbook.xml'));rels=etree.fromstring(archive.read('xl/_rels/workbook.xml.rels'))
 sheet=workbook.xpath("m:sheets/m:sheet[@name=$name]",namespaces=ns,name=sheet_name)
 if len(sheet)!=1:raise ValueError('ambiguous_target')
 rel_id=sheet[0].get('{%s}id'%ns['r']);target=rels.xpath("p:Relationship[@Id=$ident]/@Target",namespaces=ns,ident=rel_id)
 if len(target)!=1:raise ValueError('validation_failure')
 raw_target=target[0].lstrip('/')
 return raw_target if raw_target.startswith('xl/') else str(PurePosixPath('xl')/raw_target).replace('xl/../','')

def _package_set_values(source:Path,candidate:Path,changes:list[tuple[str,str,str,Any]]):
 ns='http://schemas.openxmlformats.org/spreadsheetml/2006/main';by_sheet={}
 for sheet,coordinate,action,value in changes:by_sheet.setdefault(sheet,{})[coordinate]=(action,value)
 replacements={}
 with zipfile.ZipFile(source) as archive:
  for sheet,values in by_sheet.items():
   member=_sheet_member(source,sheet);root=etree.fromstring(archive.read(member))
   for coordinate,(action,value) in values.items():
    cells=root.xpath(".//m:c[@r=$coordinate]",namespaces={'m':ns},coordinate=coordinate)
    if len(cells)!=1:raise ValueError('ambiguous_target')
    cell=cells[0]
    for child in list(cell):
     if child.tag in {f'{{{ns}}}v',f'{{{ns}}}f',f'{{{ns}}}is'}:cell.remove(child)
    cell.attrib.pop('t',None)
    if action=='formula':
     formula=etree.SubElement(cell,f'{{{ns}}}f');formula.text=value[1:];continue
    if value is None:continue
    if isinstance(value,bool):cell.set('t','b');node=etree.SubElement(cell,f'{{{ns}}}v');node.text='1' if value else '0'
    elif isinstance(value,(int,float)):node=etree.SubElement(cell,f'{{{ns}}}v');node.text=str(value)
    else:
     cell.set('t','inlineStr');inline=etree.SubElement(cell,f'{{{ns}}}is');text=etree.SubElement(inline,f'{{{ns}}}t');text.text=value
     if value[:1].isspace() or value[-1:].isspace():text.set('{http://www.w3.org/XML/1998/namespace}space','preserve')
   replacements[member]=etree.tostring(root,xml_declaration=True,encoding='UTF-8',standalone=True)
  with zipfile.ZipFile(candidate,'w') as output:
   for info in archive.infolist():output.writestr(copy.copy(info),replacements.get(info.filename,archive.read(info.filename)))

def _package_edit_cells(source:Path,candidate:Path,changes:list[tuple[str,str,str,Any]]):
 ns='http://schemas.openxmlformats.org/spreadsheetml/2006/main';styled=None;workbook=None
 try:
  roots={};members={sheet:_sheet_member(source,sheet) for sheet,_,_,_ in changes}
  with zipfile.ZipFile(source) as archive:
   for member in set(members.values()):roots[member]=etree.fromstring(archive.read(member))
  style_changes=[item for item in changes if item[2]=='style']
  style_payload=None
  if style_changes:
   fd,name=tempfile.mkstemp(prefix='.style-proof.',suffix='.xlsx',dir=candidate.parent);os.close(fd);styled=Path(name)
   workbook=load_workbook(source,data_only=False)
   for sheet,coordinate,_,style in style_changes:_apply_style(workbook[sheet][coordinate],style)
   workbook.save(styled);workbook.close();workbook=None
   with zipfile.ZipFile(styled) as generated:
    style_payload=generated.read('xl/styles.xml')
    for sheet,coordinate,_,_ in style_changes:
     member=members[sheet];generated_root=etree.fromstring(generated.read(member))
     generated_cells=generated_root.xpath(".//m:c[@r=$coordinate]",namespaces={'m':ns},coordinate=coordinate)
     original_cells=roots[member].xpath(".//m:c[@r=$coordinate]",namespaces={'m':ns},coordinate=coordinate)
     if len(generated_cells)!=1 or len(original_cells)!=1:raise ValueError('ambiguous_target')
     style_id=generated_cells[0].get('s')
     if style_id is None:original_cells[0].attrib.pop('s',None)
     else:original_cells[0].set('s',style_id)
  for sheet,coordinate,action,value in changes:
   if action=='style':continue
   member=members[sheet];cells=roots[member].xpath(".//m:c[@r=$coordinate]",namespaces={'m':ns},coordinate=coordinate)
   if len(cells)!=1:raise ValueError('ambiguous_target')
   cell=cells[0]
   for child in list(cell):
    if child.tag in {f'{{{ns}}}v',f'{{{ns}}}f',f'{{{ns}}}is'}:cell.remove(child)
   cell.attrib.pop('t',None)
   if action=='formula':
    formula=etree.SubElement(cell,f'{{{ns}}}f');formula.text=value[1:];continue
   if value is None:continue
   if isinstance(value,bool):cell.set('t','b');node=etree.SubElement(cell,f'{{{ns}}}v');node.text='1' if value else '0'
   elif isinstance(value,(int,float)):node=etree.SubElement(cell,f'{{{ns}}}v');node.text=str(value)
   else:
    cell.set('t','inlineStr');inline=etree.SubElement(cell,f'{{{ns}}}is');text=etree.SubElement(inline,f'{{{ns}}}t');text.text=value
    if value[:1].isspace() or value[-1:].isspace():text.set('{http://www.w3.org/XML/1998/namespace}space','preserve')
  replacements={member:etree.tostring(root,xml_declaration=True,encoding='UTF-8',standalone=True) for member,root in roots.items()}
  if style_payload is not None:replacements['xl/styles.xml']=style_payload
  with zipfile.ZipFile(source) as archive,zipfile.ZipFile(candidate,'w') as output:
   for info in archive.infolist():output.writestr(copy.copy(info),replacements.get(info.filename,archive.read(info.filename)))
 finally:
  if workbook is not None:workbook.close()
  if styled is not None:styled.unlink(missing_ok=True)

def _publish(output:Path,build,check):
 output.parent.mkdir(parents=True,exist_ok=True);fd,name=tempfile.mkstemp(prefix='.'+output.name+'.candidate.',suffix='.xlsx',dir=output.parent);os.close(fd);candidate=Path(name)
 try:build(candidate);check(candidate);os.replace(candidate,output)
 finally:candidate.unlink(missing_ok=True)
def _restore_members(source:Path,candidate:Path,names:set[str]):
 fd,name=tempfile.mkstemp(prefix='.'+candidate.name+'.restore.',suffix='.xlsx',dir=candidate.parent);os.close(fd);rewritten=Path(name)
 try:
  with zipfile.ZipFile(source) as original,zipfile.ZipFile(candidate) as current,zipfile.ZipFile(rewritten,'w') as output:
   original_names=set(original.namelist())
   for info in current.infolist():output.writestr(copy.copy(info),original.read(info.filename) if info.filename in names and info.filename in original_names else current.read(info.filename))
  os.replace(rewritten,candidate)
 finally:rewritten.unlink(missing_ok=True)

def _restore_outside_allowlist(source:Path,candidate:Path,allowlist:set[str]):
 fd,name=tempfile.mkstemp(prefix='.'+candidate.name+'.preserve.',suffix='.xlsx',dir=candidate.parent);os.close(fd);rewritten=Path(name)
 try:
  with zipfile.ZipFile(source) as original,zipfile.ZipFile(candidate) as current,zipfile.ZipFile(rewritten,'w') as output:
   if set(original.namelist())!=set(current.namelist()) or not allowlist<=set(current.namelist()):raise ValueError('validation_failure')
   for info in original.infolist():output.writestr(copy.copy(info),current.read(info.filename) if info.filename in allowlist else original.read(info.filename))
  os.replace(rewritten,candidate)
 finally:rewritten.unlink(missing_ok=True)

def _check_package_allowlist(source:Path,candidate:Path,allowlist:set[str]):
 with zipfile.ZipFile(source) as original,zipfile.ZipFile(candidate) as current:
  if set(original.namelist())!=set(current.namelist()):raise ValueError('validation_failure')
  for name in original.namelist():
   if name not in allowlist and original.read(name)!=current.read(name):raise ValueError('validation_failure')

def _apply_style(cell,name):
 if name not in _STYLES:raise ValueError('unsupported_style')
 for key,value in _STYLES[name].items():setattr(cell,key,copy.copy(value))

def _style_matches(cell,name):
 expected=_STYLES[name]
 if cell.number_format!=expected.get('number_format','General'):return False
 if 'font' in expected:
  expected_color=expected['font'].color;actual_color=cell.font.color
  if cell.font.bold!=expected['font'].bold or (expected_color is None)!=(actual_color is None) or expected_color is not None and (actual_color.type!=expected_color.type or actual_color.rgb!=expected_color.rgb):return False
 if 'fill' in expected:
  expected_color=expected['fill'].fgColor;actual_color=cell.fill.fgColor
  if cell.fill.fill_type!=expected['fill'].fill_type or actual_color.type!=expected_color.type or actual_color.rgb!=expected_color.rgb:return False
 return True

def validate_create_model(model:dict[str,Any])->dict[str,Any]:
 """Validate and defensively copy the closed XLSX create model."""
 if not isinstance(model,dict) or set(model) not in ({'sheets'},{'workbook_id','sheets'}):raise ValueError('validation_failure')
 if 'workbook_id' in model and (not isinstance(model['workbook_id'],str) or not _WORKBOOK_ID.fullmatch(model['workbook_id'])):raise ValueError('validation_failure')
 sheets=model['sheets']
 if not isinstance(sheets,list) or not sheets or len(sheets)>256:raise ValueError('validation_failure')
 if sum(len(spec.get('cells',{})) if isinstance(spec,dict) and isinstance(spec.get('cells'),dict) else 0 for spec in sheets)>250000:raise ValueError('unsafe_plan')
 allowed_sheet={'name','cells','freeze_panes','auto_filter','state','column_widths','row_heights','merged_ranges'};names=[]
 for spec in sheets:
  if not isinstance(spec,dict) or not {'name','cells'}<=set(spec) or set(spec)-allowed_sheet or not isinstance(spec['name'],str) or not spec['name'] or len(spec['name'])>31 or any(ch in spec['name'] for ch in '[]:*?/\\') or not isinstance(spec['cells'],dict):raise ValueError('validation_failure')
  names.append(spec['name'])
 if len(names)!=len(set(names)) or not any(spec.get('state','visible')=='visible' for spec in sheets):raise ValueError('validation_failure')
 for spec in sheets:
  if spec.get('state','visible') not in {'visible','hidden','veryHidden'}:raise ValueError('validation_failure')
  column_widths=spec.get('column_widths',{});row_heights=spec.get('row_heights',{})
  if not isinstance(column_widths,dict) or not isinstance(row_heights,dict):raise ValueError('validation_failure')
  dimensions=[*column_widths.values(),*row_heights.values()]
  if any(not isinstance(value,(int,float)) or isinstance(value,bool) or not math.isfinite(value) or value<=0 for value in dimensions):raise ValueError('validation_failure')
  for column in column_widths:
   if not isinstance(column,str) or not _COLUMN_REF.fullmatch(column):raise ValueError('validation_failure')
   if _single_cell_bounds(f'{column}1')[0]>16384:raise ValueError('validation_failure')
  for row in row_heights:
   if not isinstance(row,str) or not _ROW_REF.fullmatch(row):raise ValueError('validation_failure')
   row_number=int(row)
   if row_number<1 or row_number>1048576:raise ValueError('validation_failure')
  for coordinate,payload in spec['cells'].items():
   if not isinstance(payload,dict) or set(payload)-{'value','formula','style'} or ('value' in payload)==('formula' in payload) or payload.get('style','normal') not in _STYLES:raise ValueError('validation_failure')
   _single_cell_bounds(coordinate)
   if 'formula' in payload:
    if not _safe_formula(payload['formula']):raise ValueError('validation_failure')
   else:
    value=payload['value']
    if not _explicit_scalar(value) or isinstance(value,str) and len(value)>32767:raise ValueError('validation_failure')
  freeze_panes=spec.get('freeze_panes')
  if freeze_panes is not None:
   _single_cell_bounds(freeze_panes)
   if freeze_panes=='A1':raise ValueError('validation_failure')
  auto_filter=spec.get('auto_filter')
  if auto_filter is not None:_cell_range_bounds(auto_filter,allow_single=True)
  merged=spec.get('merged_ranges',[])
  if not isinstance(merged,list) or len(merged)>10000 or any(not isinstance(item,str) for item in merged):raise ValueError('validation_failure')
  merged_bounds=[]
  for item in merged:
   bounds=_cell_range_bounds(item,allow_single=False)
   if any(not (bounds[2]<other[0] or other[2]<bounds[0] or bounds[3]<other[1] or other[3]<bounds[1]) for other in merged_bounds):raise ValueError('validation_failure')
   merged_bounds.append(bounds)
 return copy.deepcopy(model)


def _snapshot(path:Path,view='region',sheet=None,range_ref=None):
 _admit(path);source=_sha(path);wb=load_workbook(path,data_only=False)
 if view!='region' or sheet not in wb.sheetnames or not isinstance(range_ref,str):raise ValueError('unsupported_view')
 try:min_col,min_row,max_col,max_row=range_boundaries(range_ref)
 except Exception as exc:raise ValueError('invalid_range') from exc
 if (max_row-min_row+1)*(max_col-min_col+1)>20000:raise ValueError('unsafe_plan')
 ws=wb[sheet];cells=[];rows=[]
 for row in ws.iter_rows(min_row=min_row,max_row=max_row,min_col=min_col,max_col=max_col):
  row_cells=[]
  for cell in row:
   kind='formula' if cell.data_type=='f' else 'value';value=cell.value
   if kind=='formula' and not _safe_formula(value):raise ValueError('external_reference')
   item={'id':_id(source,sheet,cell.coordinate,kind,value),'sheet':sheet,'coordinate':cell.coordinate,'row':cell.row,'column':cell.column,'kind':kind,'style_id':cell.style_id,'number_format':cell.number_format}
   item['formula' if kind=='formula' else 'value']=value;row_cells.append(item)
   if value is not None:cells.append(item)
  rows.append({'id':_row_id(source,sheet,row[0].row,row_cells),'row_number':row[0].row,'cells':row_cells})
 region_id=_region_id(source,sheet,range_ref)
 snap={'artifact_type':'xlsx','source_sha256':source,'view':'region','sheet':sheet,'range':range_ref,'region_id':region_id,'cells':cells,'rows':rows,'sheet_order':wb.sheetnames}
 snap['snapshot_sha256']=_object_sha(snap);snap['status']='ok';return snap

class XlsxArtifactTool:
 def __init__(self,workdir:Path|str):self.workdir=Path(workdir);self.workdir.mkdir(parents=True,exist_ok=True)
 def create(self,model:dict[str,Any],output:Path|str):
  output=Path(output)
  try:
   if output.suffix.lower()!='.xlsx':raise ValueError('validation_failure')
   model=validate_create_model(model)
   def build(path):
    wb=Workbook();wb.remove(wb.active)
    sheets=model['sheets']
    for spec in sheets:
     ws=wb.create_sheet(spec['name'])
     for coordinate,payload in spec.get('cells',{}).items():
      cell=ws[coordinate]
      if 'formula' in payload:
       cell.value=payload['formula']
      else:
       cell.value=payload['value']
      _apply_style(cell,payload.get('style','normal'))
     if spec.get('freeze_panes') is not None:ws.freeze_panes=spec['freeze_panes']
     if spec.get('auto_filter') is not None:ws.auto_filter.ref=spec['auto_filter']
     if spec.get('state') is not None:ws.sheet_state=spec['state']
     for column,width in spec.get('column_widths',{}).items():ws.column_dimensions[column].width=width
     for row,height in spec.get('row_heights',{}).items():ws.row_dimensions[int(row)].height=height
     for range_ref in spec.get('merged_ranges',[]):ws.merge_cells(range_ref)
    wb.save(path)
   def check_created(path):
    _candidate_openable(path);workbook=load_workbook(path,data_only=False)
    if workbook.sheetnames!=[spec['name'] for spec in model['sheets']]:raise ValueError('validation_failure')
    for spec in model['sheets']:
     sheet=workbook[spec['name']]
     if spec.get('state','visible')!=sheet.sheet_state or (spec.get('freeze_panes') or None)!=(str(sheet.freeze_panes) if sheet.freeze_panes else None) or (spec.get('auto_filter') or None)!=sheet.auto_filter.ref:raise ValueError('validation_failure')
     if sorted(spec.get('merged_ranges',[]))!=sorted(str(item) for item in sheet.merged_cells.ranges):raise ValueError('validation_failure')
     for coordinate,payload in spec['cells'].items():
      expected=payload.get('formula',payload.get('value'));cell=sheet[coordinate]
      if cell.value!=expected:raise ValueError('validation_failure')
      style=payload.get('style','normal')
      if not _style_matches(cell,style):raise ValueError('validation_failure')
     for column,width in spec.get('column_widths',{}).items():
      if sheet.column_dimensions[column].width!=width:raise ValueError('validation_failure')
     for row,height in spec.get('row_heights',{}).items():
      if sheet.row_dimensions[int(row)].height!=height:raise ValueError('validation_failure')
    workbook.close()
   _publish(output,build,check_created);return {'status':'ok','sha256':_sha(output)}
  except ValueError as exc:return _refusal(_typed_reason(exc))
  except Exception as exc:return _refusal('validation_failure',str(exc))
 def inspect(self,source:Path|str,view='region',sheet=None,range_ref=None,query=None,**kwargs):
  try:
   source=Path(source)
   if view=='inventory':
    _admit(source);result=inspect_inventory(source);result['source_sha256']=_sha(source);return result
   if view=='region':return _snapshot(source,view,sheet,range_ref)
   _admit(source);digest=_sha(source);wb=load_workbook(source,data_only=False)
   if view=='summary':
    sheets=[{'id':'tx_'+hashlib.sha256(f'{digest}|sheet|{ws.title}'.encode()).hexdigest()[:24],'name':ws.title,'state':ws.sheet_state,'used_range':ws.calculate_dimension(),'max_row':ws.max_row,'max_column':ws.max_column,'freeze_panes':str(ws.freeze_panes) if ws.freeze_panes else None,'auto_filter':ws.auto_filter.ref,'merged_ranges':[str(x) for x in ws.merged_cells.ranges],'column_widths':{key:float(value.width) for key,value in ws.column_dimensions.items() if value.width is not None},'row_heights':{str(key):float(value.height) for key,value in ws.row_dimensions.items() if value.height is not None}} for ws in wb.worksheets]
    return {'status':'ok','artifact_type':'xlsx','source_sha256':digest,'view':'summary','sheets':sheets}
   if view=='search':
    if not isinstance(query,str) or not query:raise ValueError('query_required')
    matches=[];needle=query.casefold()
    for ws in wb.worksheets:
     for row in ws.iter_rows():
      context=[cell.value for cell in row]
      for cell in row:
       if cell.value is not None and needle in str(cell.value).casefold():
        kind='formula' if cell.data_type=='f' else 'value';matches.append({'id':_id(digest,ws.title,cell.coordinate,kind,cell.value),'sheet':ws.title,'coordinate':cell.coordinate,'kind':kind,'value':cell.value,'row_context':context})
        if len(matches)>=100:return {'status':'ok','artifact_type':'xlsx','source_sha256':digest,'view':'search','query':query,'matches':matches,'truncated':True}
    return {'status':'ok','artifact_type':'xlsx','source_sha256':digest,'view':'search','query':query,'matches':matches,'truncated':False}
   raise ValueError('unsupported_view')
  except Exception as exc:return _refusal('validation_failure',str(exc))
 def plan(self,snapshot:dict[str,Any],request:dict[str,Any]):
  try:
   if snapshot.get('status')!='ok' or not isinstance(request,dict) or set(request) not in ({'operations'},{'transforms'}):raise ValueError('validation_failure')
   source_cells=[cell for row in snapshot.get('rows',[]) for cell in row.get('cells',[])];by_id={x['id']:x for x in source_cells};ops=[]
   requested=request.get('operations')
   if requested is None:
    transforms=request.get('transforms')
    if not isinstance(transforms,list) or not transforms:raise ValueError('validation_failure')
    requested=[];virtual={ident:item.get('value',item.get('formula')) for ident,item in by_id.items()}
    for transform in transforms:
     if not isinstance(transform,dict):raise ValueError('validation_failure')
     kind=transform.get('type')
     if kind=='fill_missing':
      if set(transform)!={'type','target_ids','value'} or not isinstance(transform['target_ids'],list):raise ValueError('validation_failure')
      for ident in transform['target_ids']:
       target=by_id.get(ident)
       if not target or target['kind']!='value' or virtual[ident] not in (None,''):raise ValueError('validation_failure')
       requested.append({'type':'set_cell_value','target_id':ident,'value':transform['value'],'expected_kind':'value'});virtual[ident]=transform['value']
     elif kind=='table_totals':
      if set(transform)!={'type','rows'} or not isinstance(transform['rows'],list):raise ValueError('validation_failure')
      for row in transform['rows']:
       if not isinstance(row,dict) or set(row)!={'quantity_id','unit_price_id','target_id'}:raise ValueError('validation_failure')
       qty=virtual.get(row['quantity_id']);unit=virtual.get(row['unit_price_id']);target=by_id.get(row['target_id'])
       if not isinstance(qty,(int,float)) or isinstance(qty,bool) or not isinstance(unit,(int,float)) or isinstance(unit,bool) or not target or target['kind']!='value':raise ValueError('validation_failure')
       total=qty*unit
       if not math.isfinite(total):raise ValueError('validation_failure')
       requested.append({'type':'set_cell_value','target_id':row['target_id'],'value':total,'expected_kind':'value'});virtual[row['target_id']]=total
     elif kind=='sort_rows':
      if set(transform)!={'type','region_id','keys_by_row_id','descending'} or transform['region_id']!=snapshot.get('region_id') or not isinstance(transform['keys_by_row_id'],dict) or not isinstance(transform['descending'],bool):raise ValueError('validation_failure')
      row_ids=[x['id'] for x in snapshot.get('rows',[])]
      if set(transform['keys_by_row_id'])!=set(row_ids):raise ValueError('validation_failure')
      ordered=sorted(row_ids,key=lambda ident:transform['keys_by_row_id'][ident],reverse=transform['descending']);requested.append({'type':'reorder_rows','region_id':snapshot['region_id'],'row_ids':ordered})
     else:raise ValueError('unsupported_capability')
   if not isinstance(requested,list) or not requested:raise ValueError('validation_failure')
   if len(requested)>1000:raise ValueError('unsafe_plan')
   structural_regions=[op.get('region_id') for op in requested if isinstance(op,dict) and op.get('type') in {'append_rows','reorder_rows'}]
   if len(structural_regions)!=len(set(structural_regions)):raise ValueError('conflict')
   for op in requested:
    if not isinstance(op,dict):raise ValueError('validation_failure')
    kind=op.get('type')
    if kind=='append_rows':
     if set(op)!={'type','region_id','copy_from_row_id','rows'} or op['region_id']!=snapshot.get('region_id') or not isinstance(op['rows'],list) or not op['rows'] or len(op['rows'])>10000:raise ValueError('validation_failure')
     by_row={x['id']:x for x in snapshot.get('rows',[])};template=by_row.get(op['copy_from_row_id']);width=len(snapshot['rows'][0]['cells']) if snapshot.get('rows') else 0
     if not template or not width or not all(isinstance(row,list) and len(row)==width for row in op['rows']):raise ValueError('validation_failure')
     for row in op['rows']:
      for value in row:
       if isinstance(value,dict) and (set(value)!={'formula'} or not _safe_formula(value['formula'])):raise ValueError('validation_failure')
       if not isinstance(value,dict) and (not _explicit_scalar(value) or isinstance(value,str) and len(value)>32767):raise ValueError('validation_failure')
     ops.append({'type':'append_rows','region_id':op['region_id'],'sheet':snapshot['sheet'],'range':snapshot['range'],'copy_from_row_id':op['copy_from_row_id'],'template_row':template,'rows':op['rows']});continue
    if kind=='reorder_rows':
     if set(op)!={'type','region_id','row_ids'} or op['region_id']!=snapshot.get('region_id') or not isinstance(op['row_ids'],list):raise ValueError('validation_failure')
     by_row={x['id']:x for x in snapshot.get('rows',[])}
     if len(op['row_ids'])!=len(by_row) or set(op['row_ids'])!=set(by_row):raise ValueError('validation_failure')
     ordered=[by_row[x] for x in op['row_ids']];ops.append({'type':'reorder_rows','region_id':op['region_id'],'sheet':snapshot['sheet'],'range':snapshot['range'],'row_ids':op['row_ids'],'source_rows':snapshot['rows'],'ordered_rows':ordered});continue
    if not isinstance(op,dict) or not isinstance(op.get('target_id'),str):raise ValueError('validation_failure')
    target=by_id.get(op['target_id'])
    if not target:raise ValueError('ambiguous_target')
    kind=op.get('type')
    if kind=='set_cell_value':
     if set(op)!={'type','target_id','value','expected_kind'} or op['expected_kind']!='value' or target['kind']!='value' or not _explicit_scalar(op['value']) or isinstance(op['value'],str) and len(op['value'])>32767:raise ValueError('validation_failure')
     resolved={**op}
    elif kind=='set_cell_formula':
     if set(op)!={'type','target_id','formula','expected_kind'} or op['expected_kind']!=target['kind'] or not _safe_formula(op['formula']):raise ValueError('validation_failure')
     resolved={**op}
    elif kind=='clear_cell':
     if set(op)!={'type','target_id','expected_kind'} or op['expected_kind']!=target['kind']:raise ValueError('validation_failure')
     resolved={**op}
    elif kind=='set_cell_style':
     if set(op)!={'type','target_id','style'} or op['style'] not in _STYLES:raise ValueError('validation_failure')
     resolved={**op}
    else:raise ValueError('unsupported_capability')
    resolved.update({'sheet':target['sheet'],'coordinate':target['coordinate'],'old_kind':target['kind'],'old_value':target.get('value',target.get('formula'))});ops.append(resolved)
   plan={'schema':1,'source_sha256':snapshot['source_sha256'],'snapshot_sha256':snapshot['snapshot_sha256'],'operations':ops};plan['plan_sha256']=_object_sha(plan);return {'status':'ok','plan':plan}
  except ValueError as exc:return _refusal(_typed_reason(exc))
  except Exception as exc:return _refusal('validation_failure',str(exc))
 def apply(self,source:Path|str,plan:dict[str,Any],output:Path|str):
  source=Path(source);output=Path(output);source_snapshot=None
  try:
   if output.suffix.lower()!='.xlsx':raise ValueError('validation_failure')
   if source.resolve()==output.resolve():raise ValueError('unsafe_plan')
   _admit(source)
   fd,name=tempfile.mkstemp(prefix='.source-snapshot.',suffix='.xlsx',dir=self.workdir);source_snapshot=Path(name)
   try:
    with os.fdopen(fd,'wb') as target,source.open('rb') as original:
     for block in iter(lambda:original.read(1024*1024),b''):target.write(block)
   except Exception:
    source_snapshot.unlink(missing_ok=True);raise
   _admit(source_snapshot);source=source_snapshot
   inventory=inspect_inventory(source)
   if not isinstance(plan,dict) or set(plan)!={'schema','source_sha256','snapshot_sha256','operations','plan_sha256'} or not isinstance(plan.get('operations'),list) or not plan['operations']:raise ValueError('validation_failure')
   package_preserving=all(isinstance(op,dict) and op.get('type') in PACKAGE_CELL_OPERATIONS for op in plan['operations'])
   admission=admit_operations(inventory,plan['operations'])
   if not admission['supported']:raise ValueError('unsupported_capability')
   if len(plan['operations'])>1000:raise ValueError('unsafe_plan')
   structural_regions=[op.get('region_id') for op in plan['operations'] if isinstance(op,dict) and op.get('type') in {'append_rows','reorder_rows'}]
   if len(structural_regions)!=len(set(structural_regions)):raise ValueError('conflict')
   check=dict(plan);provided=check.pop('plan_sha256')
   if provided!=_object_sha(check) or plan['source_sha256']!=_sha(source):raise ValueError('stale_snapshot')
   wb=load_workbook(source,data_only=False)
   prepared=[];seen=set();formula_changed=False;row_reorders=[];row_appends=[]
   for op in plan['operations']:
    if isinstance(op,dict) and op.get('type')=='append_rows':
     if set(op)!={'type','region_id','sheet','range','copy_from_row_id','template_row','rows'} or op['sheet'] not in wb.sheetnames or not isinstance(op['rows'],list) or not op['rows'] or op['region_id']!=_region_id(plan['source_sha256'],op['sheet'],op['range']):raise ValueError('validation_failure')
     min_col,min_row,max_col,max_row=range_boundaries(op['range']);ws=wb[op['sheet']];template=op['template_row']
     if template['id']!=op['copy_from_row_id'] or len(template['cells'])!=max_col-min_col+1 or any(not isinstance(row,list) or len(row)!=max_col-min_col+1 for row in op['rows']):raise ValueError('validation_failure')
     for row in op['rows']:
      for value in row:
       if isinstance(value,dict) and (set(value)!={'formula'} or not _safe_formula(value['formula'])):raise ValueError('validation_failure')
       if not isinstance(value,dict) and (not _explicit_scalar(value) or isinstance(value,str) and len(value)>32767):raise ValueError('validation_failure')
     for item in template['cells']:
      cell=ws[item['coordinate']];current_kind='formula' if cell.data_type=='f' else 'value'
      if current_kind!=item['kind'] or cell.value!=item.get('value',item.get('formula')) or cell.style_id!=item['style_id']:raise ValueError('stale_snapshot')
     row_appends.append((ws,min_col,max_col,max_row,template,op['rows']));formula_changed|=any(isinstance(v,dict) for row in op['rows'] for v in row);continue
    if isinstance(op,dict) and op.get('type')=='reorder_rows':
     if set(op)!={'type','region_id','sheet','range','row_ids','source_rows','ordered_rows'} or op['sheet'] not in wb.sheetnames or op['region_id']!=_region_id(plan['source_sha256'],op['sheet'],op['range']):raise ValueError('validation_failure')
     min_col,min_row,max_col,max_row=range_boundaries(op['range']);ws=wb[op['sheet']]
     source_ids=[row.get('id') for row in op['source_rows']];ordered_ids=[row.get('id') for row in op['ordered_rows']];source_by_id={row.get('id'):row for row in op['source_rows']}
     rows_bound=all(row.get('id')==_row_id(plan['source_sha256'],op['sheet'],row.get('row_number'),row.get('cells',[])) for row in op['source_rows'])
     if max_row-min_row+1!=len(op['source_rows']) or len(op['ordered_rows'])!=len(op['source_rows']) or len(source_ids)!=len(set(source_ids)) or not rows_bound or set(op['row_ids'])!=set(source_ids) or op['row_ids']!=ordered_ids or op['ordered_rows']!=[source_by_id.get(ident) for ident in op['row_ids']]:raise ValueError('validation_failure')
     for row in op['source_rows']:
      for item in row['cells']:
       cell=ws[item['coordinate']];current_kind='formula' if cell.data_type=='f' else 'value'
       if current_kind!=item['kind'] or cell.value!=item.get('value',item.get('formula')) or cell.style_id!=item['style_id']:raise ValueError('stale_snapshot')
     row_reorders.append((ws,min_col,min_row,max_col,max_row,op['ordered_rows']));continue
    required={'type','target_id','sheet','coordinate','old_kind','old_value'}
    if not isinstance(op,dict) or not required<=set(op) or op['sheet'] not in wb.sheetnames:raise ValueError('validation_failure')
    if (op['sheet'],op['coordinate']) in seen:raise ValueError('conflict')
    seen.add((op['sheet'],op['coordinate']));cell=wb[op['sheet']][op['coordinate']];current_kind='formula' if cell.data_type=='f' else 'value'
    if current_kind!=op['old_kind'] or cell.value!=op['old_value'] or _id(plan['source_sha256'],op['sheet'],op['coordinate'],current_kind,cell.value)!=op['target_id']:raise ValueError('stale_snapshot')
    kind=op['type']
    if kind=='set_cell_value':
     if set(op)!=required|{'value','expected_kind'} or op['expected_kind']!='value' or current_kind!='value' or not _explicit_scalar(op['value']) or isinstance(op['value'],str) and len(op['value'])>32767:raise ValueError('validation_failure')
     action=('value',op['value'])
    elif kind=='set_cell_formula':
     if set(op)!=required|{'formula','expected_kind'} or op['expected_kind']!=current_kind or not _safe_formula(op['formula']):raise ValueError('validation_failure')
     action=('formula',op['formula']);formula_changed=True
    elif kind=='clear_cell':
     if set(op)!=required|{'expected_kind'} or op['expected_kind']!=current_kind:raise ValueError('validation_failure')
     action=('value',None);formula_changed|=current_kind=='formula'
    elif kind=='set_cell_style':
     if set(op)!=required|{'style'} or op['style'] not in _STYLES:raise ValueError('validation_failure')
     action=('style',op['style'])
    else:raise ValueError('unsupported_capability')
    prepared.append((cell,action,op['sheet'],op['coordinate']))
   affected_sheets={sheet for _,_,sheet,_ in prepared}|{ws.title for ws,_,_,_,_,_ in row_appends}|{ws.title for ws,_,_,_,_,_ in row_reorders}
   allowlist={_sheet_member(source,sheet) for sheet in affected_sheets}
   if any(action=='style' for _,(action,_),_,_ in prepared):allowlist.add('xl/styles.xml')
   def build(path):
    if package_preserving:
     _package_edit_cells(source,path,[(sheet,coordinate,action,value) for _,(action,value),sheet,coordinate in prepared]);return
    for cell,(action,value),_,_ in prepared:
     if action in {'value','formula'}:cell.value=value
     else:_apply_style(cell,value)
    for ws,min_col,max_col,max_row,template,rows_to_add in row_appends:
     for offset,row_values in enumerate(rows_to_add,1):
      target_row=max_row+offset
      for index,(item,value) in enumerate(zip(template['cells'],row_values)):
       source_cell=ws[item['coordinate']];target=ws.cell(target_row,min_col+index)
       if source_cell.has_style:target._style=copy.copy(source_cell._style)
       target.number_format=source_cell.number_format;target.hyperlink=copy.copy(source_cell.hyperlink);target.comment=copy.copy(source_cell.comment);target.value=value['formula'] if isinstance(value,dict) else value
    for ws,min_col,min_row,max_col,max_row,ordered_rows in row_reorders:
     payload=[]
     for row in ordered_rows:
      payload.append([copy.copy(ws.cell(row['row_number'],column)) for column in range(min_col,max_col+1)])
     for target_row,cells_to_copy in enumerate(payload,min_row):
      for offset,source_cell in enumerate(cells_to_copy):
       target=ws.cell(target_row,min_col+offset);target.value=source_cell.value
       if source_cell.has_style:target._style=copy.copy(source_cell._style)
       target.number_format=source_cell.number_format;target.hyperlink=copy.copy(source_cell.hyperlink);target.comment=copy.copy(source_cell.comment)
    wb.save(path);_restore_outside_allowlist(source,path,allowlist)
   def check_candidate(candidate):
    _admit(candidate);_check_package_allowlist(source,candidate,allowlist);after=load_workbook(candidate,data_only=False)
    reordered_coordinates=set();update_values={(s,c):value for _,(action,value),s,c in prepared if action in {'value','formula'}}
    for ws,min_col,max_col,max_row,template,rows_to_add in row_appends:
     for offset,row_values in enumerate(rows_to_add,1):
      target_row=max_row+offset
      for index,value in enumerate(row_values):
       expected=value['formula'] if isinstance(value,dict) else value
       if after[ws.title].cell(target_row,min_col+index).value!=expected:raise ValueError('validation_failure')
    for ws,min_col,min_row,max_col,max_row,ordered_rows in row_reorders:
     for target_row,row in enumerate(ordered_rows,min_row):
      for offset,item in enumerate(row['cells']):
       target_coordinate=f'{get_column_letter(min_col+offset)}{target_row}';reordered_coordinates.add((ws.title,item['coordinate']));expected=update_values.get((ws.title,item['coordinate']),item.get('value',item.get('formula')))
       if after[ws.title][target_coordinate].value!=expected:raise ValueError('validation_failure')
    for _,(action,value),sheet,coordinate in prepared:
     if (sheet,coordinate) in reordered_coordinates:continue
     cell=after[sheet][coordinate]
     if action in {'value','formula'} and cell.value!=value:raise ValueError('validation_failure')
     if action=='style' and not _style_matches(cell,value):raise ValueError('validation_failure')
   _publish(output,build,check_candidate)
   return {'status':'ok','sha256':_sha(output),'changed_cells':[f'{s}!{c}' for _,_,s,c in prepared],'changed_regions':len(row_reorders)+len(row_appends),'formula_recalculation':'required' if formula_changed else 'not_required'}
  except (ValueError,KeyError,TypeError,zipfile.BadZipFile) as exc:return _refusal(_typed_reason(exc))
  except Exception as exc:return _refusal('validation_failure',str(exc))
  finally:
   if source_snapshot is not None:source_snapshot.unlink(missing_ok=True)
 def fill_template(self,source:Path|str,values:dict[str,str],output:Path|str,strict:bool=True):
  source=Path(source);output=Path(output);candidate=None;workbook=None
  try:
   if strict is not True:raise ValueError('unsupported_capability')
   if output.suffix.lower()!='.xlsx' or source.resolve()==output.resolve():raise ValueError('unsafe_plan')
   values=validate_values(values);_admit(source);digest=_sha(source);workbook=load_workbook(source,data_only=False)
   if unsupported_scope_has_marker(source):raise ValueError('unsupported_capability')
   if sum(ws.max_row*ws.max_column for ws in workbook.worksheets)>250000:raise ValueError('unsafe_plan')
   targets=[];required=set()
   for ws in workbook.worksheets:
    for row in ws.iter_rows():
     for cell in row:
      value=cell.value
      if not isinstance(value,str) or not has_marker(value):continue
      if cell.data_type=='f':raise ValueError('unsupported_capability')
      if not well_formed(value):raise ValueError('validation_failure')
      names=token_names(value)
      if not names:raise ValueError('validation_failure')
      required.update(names)
      targets.append({'id':_id(digest,ws.title,cell.coordinate,'value',value),'sheet':ws.title,'coordinate':cell.coordinate,'row':cell.row,'column':cell.column,'kind':'value','style_id':cell.style_id,'number_format':cell.number_format,'value':value})
   if not required:raise ValueError('validation_failure')
   if set(values)!=required:return _refusal('validation_failure',{'missing':sorted(required-set(values)),'unknown':sorted(set(values)-required)})
   rows=[{'id':_row_id(digest,item['sheet'],item['row'],[item]),'row_number':item['row'],'cells':[item]} for item in targets]
   snapshot={'status':'ok','artifact_type':'xlsx','source_sha256':digest,'view':'template','rows':rows};snapshot['snapshot_sha256']=_object_sha({k:v for k,v in snapshot.items() if k!='status'})
   operations=[{'type':'set_cell_value','target_id':item['id'],'value':render_tokens(item['value'],values),'expected_kind':'value'} for item in targets]
   planned=self.plan(snapshot,{'operations':operations})
   if planned.get('status')!='ok':return planned
   workbook.close();workbook=None;output.parent.mkdir(parents=True,exist_ok=True)
   fd,name=tempfile.mkstemp(prefix='.'+output.name+'.template.',suffix='.xlsx',dir=output.parent);os.close(fd);candidate=Path(name)
   applied=self.apply(source,planned['plan'],candidate)
   if applied.get('status')!='ok':return applied
   after=load_workbook(candidate,data_only=False)
   try:
    for ws in after.worksheets:
     if sum(1 for row in ws.iter_rows() for cell in row if isinstance(cell.value,str) and has_marker(cell.value)):raise ValueError('validation_failure')
   finally:after.close()
   if package_has_marker(candidate):raise ValueError('validation_failure')
   os.replace(candidate,output);candidate=None;applied['output']=str(output);applied['template']={'strict':True,'keys':sorted(required),'resolved_targets':len(targets)};return applied
  except ValueError as exc:return _refusal(_typed_reason(exc))
  except (KeyError,TypeError,zipfile.BadZipFile):return _refusal('validation_failure')
  finally:
   if workbook is not None:workbook.close()
   if candidate is not None:candidate.unlink(missing_ok=True)
 def validate(self,source:Path|str,before:Path|str|None=None):
  source=Path(source)
  try:
   _admit(source);load_workbook(source,data_only=False)
   changed=[]
   if before:
    with zipfile.ZipFile(before) as a,zipfile.ZipFile(source) as b:
     ah={x.filename:hashlib.sha256(a.read(x.filename)).hexdigest() for x in a.infolist()};bh={x.filename:hashlib.sha256(b.read(x.filename)).hexdigest() for x in b.infolist()};changed=sorted(k for k in ah|bh if ah.get(k)!=bh.get(k))
   unexpected=[x for x in changed if not x.startswith('xl/worksheets/') and x!='xl/styles.xml']
   return {'status':'valid' if not unexpected else 'invalid','package_valid':True,'changed_members':changed,'unexpected_changed_members':unexpected,'application_compatibility':'not_executed','formula_recalculation':'not_executed','visual_fidelity':'not_executed'}
  except Exception:return {'status':'invalid','error':'invalid_package'}
