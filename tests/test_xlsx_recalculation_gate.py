from __future__ import annotations

import argparse
import copy
import hashlib
import importlib.util
import json
import shutil
import subprocess
import sys
import tempfile
import textwrap
import unittest
import zipfile
from pathlib import Path
from unittest import mock

from lxml import etree
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "application_gates" / "verify_xlsx_recalculation.py"
ROUNDTRIP = ROOT / "scripts" / "application_gates" / "run_xlsx_recalculation.py"
NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


class XlsxRecalculationGateTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.source = self.root / "source.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Calc"
        sheet["A1"] = 1
        sheet["B1"] = "=A1+1"
        sheet["C1"] = "=_xlfn.UNKNOWN(A1)"
        workbook.save(self.source)

    def tearDown(self) -> None:
        self.temp.cleanup()

    def with_cached_values(self, values: dict[str, str]) -> Path:
        normalized = self.root / "normalized.xlsx"
        with zipfile.ZipFile(self.source) as original, zipfile.ZipFile(normalized, "w") as target:
            for info in original.infolist():
                payload = original.read(info.filename)
                if info.filename == "xl/worksheets/sheet1.xml":
                    root = etree.fromstring(payload)
                    for cell_ref, value in values.items():
                        cell = root.xpath(
                            ".//m:c[@r=$cell_ref]",
                            namespaces={"m": NS},
                            cell_ref=cell_ref,
                        )[0]
                        cached = cell.find(f"{{{NS}}}v")
                        if cached is None:
                            cached = etree.SubElement(cell, f"{{{NS}}}v")
                        cached.text = value
                    payload = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)
                target.writestr(copy.copy(info), payload)
        return normalized

    def run_gate(self, normalized: Path, cases: dict[str, object]) -> tuple[subprocess.CompletedProcess[str], dict]:
        case_path = self.root / "cases.json"
        case_path.write_text(json.dumps(cases))
        completed = subprocess.run(
            [sys.executable, str(SCRIPT), "--source", str(self.source), "--normalized", str(normalized), "--cases", str(case_path)],
            text=True,
            capture_output=True,
            check=False,
        )
        return completed, json.loads(completed.stdout)

    def test_declared_cached_number_is_verified_with_tolerance(self) -> None:
        normalized = self.with_cached_values({"B1": "2.0004"})
        completed, report = self.run_gate(
            normalized,
            {
                "schema_version": 1,
                "cells": [
                    {
                        "sheet": "Calc",
                        "cell": "B1",
                        "mode": "verify",
                        "expected_type": "number",
                        "expected_value": 2.0,
                        "tolerance": 0.001,
                    }
                ],
            },
        )
        self.assertEqual(completed.returncode, 0, completed.stderr)
        self.assertEqual(report["status"], "ok")
        self.assertTrue(report["recalculation_requested"])
        self.assertTrue(report["cached_values_verified"])
        self.assertEqual(report["verified_cells"], ["Calc!B1"])
        self.assertEqual(report["unsupported_formula"], [])

    def test_unsupported_formula_is_classified_without_silent_pass(self) -> None:
        normalized = self.with_cached_values({"C1": "#NAME?"})
        completed, report = self.run_gate(
            normalized,
            {
                "schema_version": 1,
                "cells": [
                    {"sheet": "Calc", "cell": "C1", "mode": "unsupported"}
                ],
            },
        )
        self.assertEqual(completed.returncode, 0, completed.stderr)
        self.assertEqual(report["status"], "ok")
        self.assertFalse(report["cached_values_verified"])
        self.assertEqual(report["verified_cells"], [])
        self.assertEqual(report["unsupported_formula"], ["Calc!C1"])

    def test_wrong_cache_and_malformed_contract_fail_closed(self) -> None:
        normalized = self.with_cached_values({"B1": "3"})
        completed, report = self.run_gate(
            normalized,
            {
                "schema_version": 1,
                "cells": [
                    {
                        "sheet": "Calc",
                        "cell": "B1",
                        "mode": "verify",
                        "expected_type": "number",
                        "expected_value": 2,
                        "tolerance": 0,
                    }
                ],
            },
        )
        self.assertEqual(completed.returncode, 2)
        self.assertEqual(report["status"], "refused")
        self.assertEqual(report["reason"], "cached_value_mismatch")

        malformed, malformed_report = self.run_gate(normalized, {"schema_version": 1, "cells": []})
        self.assertEqual(malformed.returncode, 2)
        self.assertEqual(malformed_report, {"schema_version": 1, "status": "refused", "reason": "invalid_contract"})

    def test_normalized_formula_must_match_source_formula(self) -> None:
        normalized = self.with_cached_values({"B1": "2"})
        rewritten = self.root / "formula-destroyed.xlsx"
        with zipfile.ZipFile(normalized) as original, zipfile.ZipFile(rewritten, "w") as target:
            for info in original.infolist():
                payload = original.read(info.filename)
                if info.filename == "xl/worksheets/sheet1.xml":
                    root = etree.fromstring(payload)
                    cell = root.xpath('.//m:c[@r="B1"]', namespaces={"m": NS})[0]
                    formula = cell.find(f"{{{NS}}}f")
                    cell.remove(formula)
                    payload = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)
                target.writestr(copy.copy(info), payload)
        completed, report = self.run_gate(
            rewritten,
            {
                "schema_version": 1,
                "cells": [
                    {
                        "sheet": "Calc",
                        "cell": "B1",
                        "mode": "verify",
                        "expected_type": "number",
                        "expected_value": 2,
                        "tolerance": 0,
                    }
                ],
            },
        )
        self.assertEqual(completed.returncode, 2)
        self.assertEqual(report["reason"], "normalized_formula_mismatch")

    def test_formula_and_cache_reads_share_one_immutable_snapshot(self) -> None:
        normalized = self.with_cached_values({"B1": "2"})
        cases = self.root / "snapshot-cases.json"
        cases.write_text(
            json.dumps(
                {
                    "schema_version": 1,
                    "cells": [
                        {
                            "sheet": "Calc",
                            "cell": "B1",
                            "mode": "verify",
                            "expected_type": "number",
                            "expected_value": 2,
                            "tolerance": 0,
                        }
                    ],
                }
            )
        )
        spec = importlib.util.spec_from_file_location("xlsx_verify_snapshot", SCRIPT)
        assert spec is not None and spec.loader is not None
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        expected_hash = hashlib.sha256(normalized.read_bytes()).hexdigest()
        real_load = module.load_workbook
        opened: list[Path] = []

        def swap_original_after_formula_open(path, *args, **kwargs):
            opened.append(Path(path))
            workbook = real_load(path, *args, **kwargs)
            if len(opened) == 2:
                with zipfile.ZipFile(normalized, "a") as archive:
                    archive.writestr("race-marker", b"changed")
            return workbook

        with mock.patch.object(module, "load_workbook", side_effect=swap_original_after_formula_open):
            report = module.verify(self.source, normalized, cases)
        self.assertEqual(report["status"], "ok")
        self.assertEqual(report["normalized_snapshot_sha256"], expected_hash)
        self.assertEqual(opened[1], opened[2])
        self.assertNotEqual(opened[1], normalized)

    def test_cleanup_directory_swap_cannot_return_success(self) -> None:
        spec = importlib.util.spec_from_file_location("review_run_xlsx_recalculation", ROUNDTRIP)
        module = importlib.util.module_from_spec(spec)
        sys.path.insert(0, str(ROUNDTRIP.parent))
        try:
            spec.loader.exec_module(module)
        finally:
            sys.path.pop(0)
        cases = self.root / "cleanup-cases.json"
        cases.write_text(
            json.dumps(
                {
                    "schema_version": 1,
                    "cells": [
                        {
                            "sheet": "Calc",
                            "cell": "B1",
                            "mode": "verify",
                            "expected_type": "number",
                            "expected_value": 2,
                            "tolerance": 0,
                        }
                    ],
                }
            )
        )
        normalized = self.with_cached_values({"B1": "2"})
        workdir = self.root / "cleanup-work"

        def fake_process(argv, **_kwargs):
            output = Path(argv[argv.index("--outdir") + 1]) / "artifact.xlsx"
            shutil.copyfile(normalized, output)
            return 0

        real_rmtree = module.shutil.rmtree
        swapped = False

        def swap_then_remove(path, *args, **kwargs):
            nonlocal swapped
            current = Path(path)
            if not swapped and current.name.startswith(".xlsx-recalculation."):
                swapped = True
                saved_root = workdir.with_name(workdir.name + ".saved")
                workdir.rename(saved_root)
                workdir.mkdir(mode=0o700)
            return real_rmtree(path, *args, **kwargs)

        args = argparse.Namespace(
            source=str(self.source),
            cases=str(cases),
            workdir=str(workdir),
            executable=sys.executable,
            runtime_version="LibreOffice Fake 1.0",
            runtime_image_digest="sha256:" + "c" * 64,
            timeout_seconds=5,
        )
        with mock.patch.object(module, "_run_process", side_effect=fake_process), mock.patch.object(
            module.shutil, "rmtree", side_effect=swap_then_remove
        ), self.assertRaises(module.GateRefusal) as raised:
            module.run(args)
        self.assertEqual(raised.exception.reason, "cleanup_failure")

    def test_roundtrip_runner_uses_fixed_application_boundary_then_verifies_cache(self) -> None:
        executable = self.root / "fake-calc"
        executable.write_text(
            textwrap.dedent(
                f"""\
                #!{sys.executable}
                import copy, pathlib, sys, zipfile
                from lxml import etree
                source = pathlib.Path(sys.argv[-1])
                output = pathlib.Path(sys.argv[sys.argv.index('--outdir') + 1]) / source.name
                output.parent.mkdir(parents=True, exist_ok=True)
                namespace = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                with zipfile.ZipFile(source) as original, zipfile.ZipFile(output, 'w') as target:
                    for info in original.infolist():
                        payload = original.read(info.filename)
                        if info.filename == 'xl/worksheets/sheet1.xml':
                            root = etree.fromstring(payload)
                            cell = root.xpath('.//m:c[@r="B1"]', namespaces={{'m': namespace}})[0]
                            cached = cell.find('{{%s}}v' % namespace)
                            if cached is None:
                                cached = etree.SubElement(cell, '{{%s}}v' % namespace)
                            cached.text = '2'
                            payload = etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)
                        target.writestr(copy.copy(info), payload)
                """
            )
        )
        executable.chmod(executable.stat().st_mode | 0o100)
        cases = self.root / "roundtrip-cases.json"
        cases.write_text(
            json.dumps(
                {
                    "schema_version": 1,
                    "cells": [
                        {
                            "sheet": "Calc",
                            "cell": "B1",
                            "mode": "verify",
                            "expected_type": "number",
                            "expected_value": 2,
                            "tolerance": 0,
                        }
                    ],
                }
            )
        )
        workdir = self.root / "roundtrip-work"
        completed = subprocess.run(
            [
                sys.executable,
                str(ROUNDTRIP),
                "--source",
                str(self.source),
                "--cases",
                str(cases),
                "--workdir",
                str(workdir),
                "--executable",
                str(executable),
                "--runtime-version",
                "LibreOffice Fake 1.0",
                "--runtime-image-digest",
                "sha256:" + "c" * 64,
            ],
            text=True,
            capture_output=True,
            check=False,
        )
        report = json.loads(completed.stdout)
        self.assertEqual(completed.returncode, 0, f"{completed.stderr}\n{report}")
        self.assertEqual(report["status"], "ok")
        self.assertTrue(report["cached_values_verified"])
        self.assertEqual(report["runtime_identity"]["image_digest"], "sha256:" + "c" * 64)
        self.assertEqual(list(workdir.iterdir()), [])


if __name__ == "__main__":
    unittest.main()
