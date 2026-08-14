"""Verify declared cached XLSX values after an application round-trip.

This gate never infers recalculation from a process exit. It consumes a source
workbook, an application-normalized workbook, and a closed list of cells whose
cached values must be checked or explicitly classified as unsupported.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import stat
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_MAX_XLSX_BYTES = 256 * 1024 * 1024
_MAX_CASES_BYTES = 1024 * 1024
_CELL = re.compile(r"\$?[A-Z]{1,3}\$?[1-9][0-9]{0,6}\Z")
_ALLOWED_TYPES = {"number", "string", "boolean", "error", "blank"}


class GateRefusal(ValueError):
    def __init__(self, reason: str):
        super().__init__(reason)
        self.reason = reason


def _regular_file(path: Path, maximum_bytes: int, suffix: str) -> None:
    info = path.lstat()
    if (
        stat.S_ISLNK(info.st_mode)
        or not stat.S_ISREG(info.st_mode)
        or not 0 < info.st_size <= maximum_bytes
        or path.suffix.lower() != suffix
    ):
        raise GateRefusal("invalid_input")


def _load_contract(path: Path) -> list[dict[str, Any]]:
    _regular_file(path, _MAX_CASES_BYTES, ".json")
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict) or set(payload) != {"schema_version", "cells"}:
        raise GateRefusal("invalid_contract")
    cells = payload.get("cells")
    if payload.get("schema_version") != 1 or not isinstance(cells, list) or not 1 <= len(cells) <= 1000:
        raise GateRefusal("invalid_contract")
    seen: set[tuple[str, str]] = set()
    validated: list[dict[str, Any]] = []
    for item in cells:
        if not isinstance(item, dict) or item.get("mode") not in {"verify", "unsupported"}:
            raise GateRefusal("invalid_contract")
        sheet = item.get("sheet")
        coordinate = item.get("cell")
        if (
            not isinstance(sheet, str)
            or not 1 <= len(sheet) <= 31
            or any(character in sheet for character in "[]:*?/\\")
            or not isinstance(coordinate, str)
            or not _CELL.fullmatch(coordinate)
            or (sheet, coordinate) in seen
        ):
            raise GateRefusal("invalid_contract")
        seen.add((sheet, coordinate))
        if item["mode"] == "unsupported":
            if set(item) != {"sheet", "cell", "mode"}:
                raise GateRefusal("invalid_contract")
        else:
            if set(item) != {"sheet", "cell", "mode", "expected_type", "expected_value", "tolerance"}:
                raise GateRefusal("invalid_contract")
            expected_type = item.get("expected_type")
            tolerance = item.get("tolerance")
            expected = item.get("expected_value")
            if expected_type not in _ALLOWED_TYPES or not isinstance(tolerance, (int, float)) or isinstance(tolerance, bool):
                raise GateRefusal("invalid_contract")
            if not math.isfinite(float(tolerance)) or tolerance < 0:
                raise GateRefusal("invalid_contract")
            if expected_type == "number" and (
                not isinstance(expected, (int, float)) or isinstance(expected, bool) or not math.isfinite(float(expected))
            ):
                raise GateRefusal("invalid_contract")
            if expected_type == "string" and not isinstance(expected, str):
                raise GateRefusal("invalid_contract")
            if expected_type == "boolean" and not isinstance(expected, bool):
                raise GateRefusal("invalid_contract")
            if expected_type == "error" and (not isinstance(expected, str) or not expected.startswith("#")):
                raise GateRefusal("invalid_contract")
            if expected_type == "blank" and expected is not None:
                raise GateRefusal("invalid_contract")
            if expected_type != "number" and tolerance != 0:
                raise GateRefusal("invalid_contract")
        validated.append(dict(item))
    return validated


def _matches(value: Any, expected_type: str, expected: Any, tolerance: float) -> bool:
    if expected_type == "number":
        return isinstance(value, (int, float)) and not isinstance(value, bool) and abs(float(value) - float(expected)) <= tolerance
    if expected_type == "string":
        return isinstance(value, str) and not value.startswith("#") and value == expected
    if expected_type == "boolean":
        return isinstance(value, bool) and value is expected
    if expected_type == "error":
        return isinstance(value, str) and value == expected
    return value is None


def _snapshot_xlsx(path: Path, directory: Path) -> tuple[Path, str, int]:
    _regular_file(path, _MAX_XLSX_BYTES, ".xlsx")
    flags = os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0) | getattr(os, "O_CLOEXEC", 0)
    descriptor = os.open(path, flags)
    target = directory / "normalized.xlsx"
    digest = hashlib.sha256()
    size = 0
    try:
        before = os.fstat(descriptor)
        target_fd = os.open(
            target,
            os.O_WRONLY | os.O_CREAT | os.O_EXCL | getattr(os, "O_NOFOLLOW", 0),
            0o600,
        )
        with os.fdopen(descriptor, "rb", closefd=True) as source, os.fdopen(target_fd, "wb") as output:
            descriptor = -1
            for block in iter(lambda: source.read(1024 * 1024), b""):
                size += len(block)
                if size > _MAX_XLSX_BYTES:
                    raise GateRefusal("invalid_input")
                digest.update(block)
                output.write(block)
            after = os.fstat(source.fileno())
        current = path.lstat()
    finally:
        if descriptor >= 0:
            os.close(descriptor)
    fields = ("st_dev", "st_ino", "st_mode", "st_size", "st_mtime_ns", "st_ctime_ns")
    expected = tuple(getattr(before, field) for field in fields)
    if tuple(getattr(after, field) for field in fields) != expected or tuple(
        getattr(current, field) for field in fields
    ) != expected:
        raise GateRefusal("stale_snapshot")
    return target, digest.hexdigest(), size


def verify(source: Path, normalized: Path, cases_path: Path) -> dict[str, Any]:
    _regular_file(source, _MAX_XLSX_BYTES, ".xlsx")
    cases = _load_contract(cases_path)
    with tempfile.TemporaryDirectory(prefix="xlsx-recalculation-verify.") as temp:
        snapshot, normalized_sha256, normalized_bytes = _snapshot_xlsx(normalized, Path(temp))
        formulas = load_workbook(source, data_only=False, read_only=True)
        normalized_formulas = load_workbook(snapshot, data_only=False, read_only=True)
        cached = load_workbook(snapshot, data_only=True, read_only=True)
        verified: list[str] = []
        unsupported: list[str] = []
        try:
            for item in cases:
                sheet = item["sheet"]
                coordinate = item["cell"]
                label = f"{sheet}!{coordinate}"
                if sheet not in formulas.sheetnames or sheet not in normalized_formulas.sheetnames or sheet not in cached.sheetnames:
                    raise GateRefusal("invalid_contract")
                formula = formulas[sheet][coordinate].value
                if not isinstance(formula, str) or not formula.startswith("="):
                    raise GateRefusal("invalid_contract")
                if normalized_formulas[sheet][coordinate].value != formula:
                    raise GateRefusal("normalized_formula_mismatch")
                if item["mode"] == "unsupported":
                    unsupported.append(label)
                    continue
                value = cached[sheet][coordinate].value
                if not _matches(value, item["expected_type"], item["expected_value"], float(item["tolerance"])):
                    raise GateRefusal("cached_value_mismatch")
                verified.append(label)
        finally:
            formulas.close()
            normalized_formulas.close()
            cached.close()
    return {
        "schema_version": 1,
        "status": "ok",
        "recalculation_requested": True,
        "cached_values_verified": bool(verified),
        "verified_cells": verified,
        "unsupported_formula": unsupported,
        "normalized_snapshot_sha256": normalized_sha256,
        "normalized_snapshot_bytes": normalized_bytes,
        "process_exit_alone_is_evidence": False,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True)
    parser.add_argument("--normalized", required=True)
    parser.add_argument("--cases", required=True)
    args = parser.parse_args()
    try:
        report = verify(Path(args.source), Path(args.normalized), Path(args.cases))
    except GateRefusal as exc:
        report = {"schema_version": 1, "status": "refused", "reason": exc.reason}
    except (OSError, ValueError, TypeError, KeyError, json.JSONDecodeError):
        report = {"schema_version": 1, "status": "refused", "reason": "invalid_input"}
    print(json.dumps(report, ensure_ascii=False, separators=(",", ":")))
    return 0 if report["status"] == "ok" else 2


if __name__ == "__main__":
    raise SystemExit(main())
